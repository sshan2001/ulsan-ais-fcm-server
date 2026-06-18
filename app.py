import json
import math
import os
import time
import requests
import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from flask import Flask, jsonify, request
from flask_cors import CORS

try:
    import firebase_admin
    from firebase_admin import credentials, messaging
except Exception:
    firebase_admin = None
    credentials = None
    messaging = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

app = Flask(__name__)
CORS(app)

DATA_DIR = Path(os.environ.get("DATA_DIR", "/tmp/ulsan_fcm_server"))
DATA_DIR.mkdir(parents=True, exist_ok=True)

TOKENS_FILE = DATA_DIR / "tokens.json"
WATCH_FILE = DATA_DIR / "watchlist.json"
EVENT_FILE = DATA_DIR / "event_history.json"
AIS_ALERT_FILE = DATA_DIR / "ais_alert_state.json"
AIS_SHIP_STATE_FILE = DATA_DIR / "ais_ship_state.json"
AIS_BASELINE_FILE = DATA_DIR / "ais_baseline_pending.json"
PORTMIS_FILE = DATA_DIR / "portmis_weekly.json"
PORTMIS_BACKUP_FILE = DATA_DIR / "portmis_weekly_last_upload.json"


AIS_ALERT_COOLDOWN_MINUTES = int(os.environ.get("AIS_ALERT_COOLDOWN_MINUTES", "30"))
AUTO_CHECK_ENABLED = os.environ.get("AUTO_CHECK_ENABLED", "true").strip().lower() in ("1", "true", "yes", "y", "on")
AUTO_CHECK_INTERVAL_SECONDS = int(os.environ.get("AUTO_CHECK_INTERVAL_SECONDS", "60"))

SERVER_VERSION = "3.1.9-portmis-header-parser"
SERVER_STARTED_AT = datetime.now(timezone.utc).isoformat()

# 3.1.2 서버 상태 진단/자동감시 워치독 설정
# Render 유료 전환 후에도 감시 루프가 실제로 계속 도는지 앱/브라우저에서 확인하기 위한 값입니다.
# 서버 상태 API: GET /server-health?api_key=...
SERVER_HEALTH_FILE = DATA_DIR / "server_health.json"
WATCHDOG_STALE_SECONDS = int(os.environ.get(
    "WATCHDOG_STALE_SECONDS",
    str(max(180, AUTO_CHECK_INTERVAL_SECONDS * 2 + 60)),
))

# 3.1.1 항로 좌표 판정/이벤트 흐름 설정
# 일반 이벤트는 AIS_ALERT_COOLDOWN_MINUTES(기본 30분) 쿨다운을 유지합니다.
# 아래 특수 이벤트는 30분 쿨다운을 무시하고,
# 선박 + 구역 + 단계 기준으로 1회씩 발송합니다.
SPECIAL_EVENT_PREFIXES = (
    "berth_approaching",
    "berth_completed",
    "anchorage_approaching",
    "anchorage_completed",
)

SERVER_API_KEY = os.environ.get("SERVER_API_KEY", "").strip()
FIREBASE_SERVICE_ACCOUNT_JSON = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
FIREBASE_SERVICE_ACCOUNT_FILE = os.environ.get("FIREBASE_SERVICE_ACCOUNT_FILE", "").strip()

firebase_ready = False
firebase_error = ""

auto_checker_started = False
auto_checker_pid = 0
auto_checker_thread = None
auto_checker_last_run = ""
auto_checker_last_started = ""
auto_checker_last_completed = ""
auto_checker_last_success = ""
auto_checker_last_error = ""
auto_checker_last_result: Dict[str, Any] = {}
auto_checker_run_count = 0
auto_checker_consecutive_errors = 0
last_alert_sent_at = ""
last_ais_count = 0
last_detected_count = 0
last_event_count = 0
last_sent_count = 0
last_watched_count = 0
last_check_duration_ms = 0
health_lock = threading.Lock()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def require_api_key() -> bool:
    if not SERVER_API_KEY:
        return True
    supplied = request.headers.get("X-API-Key", "").strip() or request.args.get("api_key", "").strip()
    return supplied == SERVER_API_KEY


def read_json(path: Path, fallback: Any) -> Any:
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return fallback


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_iso_time(value: str):
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


def seconds_since_iso(value: str):
    dt = parse_iso_time(value) if value else None
    if dt is None:
        return None
    try:
        return int((datetime.now(timezone.utc) - dt).total_seconds())
    except Exception:
        return None




def is_no_watched_result(result: Any) -> bool:
    """감시 선박이 없는 상태는 서버 오류가 아니라 대기/정상 상태로 처리합니다."""
    return isinstance(result, dict) and str(result.get("error") or "").strip() == "no watched ships"

def safe_len(value: Any) -> int:
    try:
        return len(value)
    except Exception:
        return 0


def mark_baseline_pending_for_missing_state(ships: List[str], source: str = "register-watch", token: str = "") -> List[str]:
    """
    앱이 추적 선박을 새로 등록/재동기화한 직후에는
    이미 접안/투묘 중인 선박의 현재 상태를 먼저 기준값으로 저장합니다.

    목적:
    - Render 재시작/재배포 후 앱이 감시목록을 다시 올렸을 때
      AIS 최초감지 + 접안완료 + 투묘완료 알림이 한꺼번에 쏟아지는 현상 방지
    - 서버에 이전 선박 상태가 없는 선박만 1회 baseline 대상으로 표시
    """
    normalized = unique_ship_list(ships)
    if not normalized:
        return []

    state = read_json(AIS_SHIP_STATE_FILE, {})
    if not isinstance(state, dict):
        state = {}

    pending = read_json(AIS_BASELINE_FILE, {})
    if not isinstance(pending, dict):
        pending = {}

    added: List[str] = []
    ts = now_iso()
    token_hint = str(token or "")[:12]

    for ship in normalized:
        name = normalize_ship_name(ship)
        if not name:
            continue
        if isinstance(state.get(name), dict):
            continue
        if name not in pending:
            pending[name] = {
                "shipName": name,
                "createdAt": ts,
                "source": source,
                "tokenHint": token_hint,
                "reason": "initial baseline pending",
            }
            added.append(name)

    if added:
        write_json(AIS_BASELINE_FILE, pending)

    return added


def prune_baseline_pending_to_watched(watched_ships: set[str]) -> None:
    pending = read_json(AIS_BASELINE_FILE, {})
    if not isinstance(pending, dict) or not pending:
        return
    keep = {k: v for k, v in pending.items() if normalize_ship_name(k) in watched_ships}
    if len(keep) != len(pending):
        write_json(AIS_BASELINE_FILE, keep)


def is_special_event_type(event_type: str) -> bool:
    return any(str(event_type).startswith(prefix) for prefix in SPECIAL_EVENT_PREFIXES)


def should_send_ais_alert(ship_name: str, event_type: str, force: bool = False) -> tuple[bool, str]:
    """
    일반 이벤트: 같은 선박 + 같은 이벤트는 AIS_ALERT_COOLDOWN_MINUTES 동안 차단합니다.
    특수 이벤트: berth_approaching/completed, anchorage_approaching/completed 계열은
    30분 쿨다운을 무시하고 정확히 같은 event_type 기준 1회만 발송합니다.

    특수 이벤트의 event_type은 반드시 구역을 포함해야 합니다.
    예: berth_approaching:ULSAN_PORT_BERTH_AREA
        berth_completed:SK_BUOY
        anchorage_completed:M3
    """
    if force:
        return True, "forced"

    normalized_ship = normalize_ship_name(ship_name)
    event_type = str(event_type).strip()
    key = f"{normalized_ship}::{event_type}"
    state = read_json(AIS_ALERT_FILE, {})
    now = datetime.now(timezone.utc)

    previous_raw = state.get(key, {}).get("sentAt") if isinstance(state.get(key), dict) else None
    previous_dt = parse_iso_time(previous_raw) if previous_raw else None

    # 특수 이벤트는 30분 쿨다운 대상이 아닙니다.
    # 대신 같은 선박 + 같은 구역 + 같은 단계는 반복 발송하지 않습니다.
    if is_special_event_type(event_type):
        if previous_dt is not None:
            return False, "special event already sent"

        state[key] = {
            "shipName": normalized_ship,
            "eventType": event_type,
            "special": True,
            "sentAt": now.isoformat(),
        }
        write_json(AIS_ALERT_FILE, cleanup_alert_state(state, now))
        return True, "ok special"

    if previous_dt is not None:
        diff_minutes = (now - previous_dt).total_seconds() / 60
        if diff_minutes < AIS_ALERT_COOLDOWN_MINUTES:
            remain = max(1, int(AIS_ALERT_COOLDOWN_MINUTES - diff_minutes))
            return False, f"cooldown {remain}min remaining"

    state[key] = {
        "shipName": normalized_ship,
        "eventType": event_type,
        "special": False,
        "sentAt": now.isoformat(),
    }

    write_json(AIS_ALERT_FILE, cleanup_alert_state(state, now))
    return True, "ok"


def cleanup_alert_state(state: Dict[str, Any], now: datetime) -> Dict[str, Any]:
    """
    중복방지 기록 정리.
    - 일반 이벤트: 24시간 보관
    - 특수 이벤트: 48시간 보관
      같은 부두/묘지에서 장시간 0.0kn 상태가 유지될 때 반복 알림을 막기 위함입니다.
    """
    cleaned = {}
    for item_key, item in state.items():
        if not isinstance(item, dict):
            continue
        sent_at = parse_iso_time(item.get("sentAt"))
        if sent_at is None:
            continue
        keep_seconds = 48 * 3600 if item.get("special") else 24 * 3600
        if (now - sent_at).total_seconds() <= keep_seconds:
            cleaned[item_key] = item
    return cleaned

def normalize_ship_name(value: Any) -> str:
    return str(value).strip().upper()


def unique_ship_list(values: Any) -> List[str]:
    if not isinstance(values, list):
        return []
    result: List[str] = []
    seen = set()
    for value in values:
        name = normalize_ship_name(value)
        if not name or name in seen:
            continue
        seen.add(name)
        result.append(name)
    return result


def unique_token_list(values: Any) -> List[str]:
    if not isinstance(values, list):
        return []
    result: List[str] = []
    seen = set()
    for value in values:
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def notification_off_ship_list_from_payload(payload: Dict[str, Any]) -> List[str]:
    """
    앱 버전에 따라 선박별 알림 OFF 목록 키 이름이 달라질 수 있어 여러 이름을 허용합니다.
    알림 OFF는 추적은 유지하되, 해당 휴대폰 토큰으로 FCM 푸시만 보내지 않기 위한 설정입니다.
    """
    for key in (
        "notificationOffShips",
        "mutedShips",
        "alertOffShips",
        "notificationMutedShips",
        "shipNotificationOff",
    ):
        values = payload.get(key)
        if isinstance(values, list):
            return unique_ship_list(values)

    settings = payload.get("notificationSettings")
    if isinstance(settings, dict):
        off = []
        for ship_name, enabled in settings.items():
            if enabled is False:
                off.append(ship_name)
        return unique_ship_list(off)

    return []


def watch_summary(watch: Dict[str, Any], token: str = "", ship_name: str = "") -> Dict[str, Any]:
    total_watch_devices = len(watch)
    total_watch_links = 0
    total_watch_ships_set = set()
    my_watch_count = 0
    my_notification_off_count = 0
    ship_watch_device_count = 0
    ship_notification_enabled_device_count = 0
    normalized_ship = normalize_ship_name(ship_name)

    for device_token, item in watch.items():
        ships = unique_ship_list(item.get("ships", [])) if isinstance(item, dict) else []
        notification_off_ships = unique_ship_list(item.get("notificationOffShips", [])) if isinstance(item, dict) else []
        total_watch_links += len(ships)
        total_watch_ships_set.update(ships)

        if token and device_token == token:
            my_watch_count = len(ships)
            my_notification_off_count = len(notification_off_ships)

        if normalized_ship and normalized_ship in ships:
            ship_watch_device_count += 1
            if normalized_ship not in notification_off_ships:
                ship_notification_enabled_device_count += 1

    return {
        "watchDeviceCount": total_watch_devices,
        "totalWatchShips": len(total_watch_ships_set),
        "totalWatchLinks": total_watch_links,
        "myWatchCount": my_watch_count,
        "myNotificationOffCount": my_notification_off_count,
        "shipWatchDeviceCount": ship_watch_device_count,
        "shipNotificationEnabledDeviceCount": ship_notification_enabled_device_count,
    }


def update_health_from_check_result(result: Dict[str, Any], duration_ms: int = 0) -> None:
    """자동 AIS 검사 결과를 서버 진단값으로 저장합니다."""
    global auto_checker_last_completed, auto_checker_last_success, auto_checker_last_error
    global auto_checker_consecutive_errors, last_ais_count, last_detected_count
    global last_event_count, last_sent_count, last_watched_count, last_check_duration_ms
    global last_alert_sent_at

    completed_at = now_iso()
    with health_lock:
        auto_checker_last_completed = completed_at
        last_check_duration_ms = int(duration_ms or 0)

        if isinstance(result, dict):
            last_watched_count = int(result.get("watchedCount") or 0)
            last_ais_count = int(result.get("aisCount") or 0)
            last_detected_count = int(result.get("detectedCount") or 0)
            last_event_count = int(result.get("eventCount") or 0)
            sent_ships = result.get("sentShips") if isinstance(result.get("sentShips"), list) else []
            last_sent_count = len(sent_ships)

            if last_sent_count > 0:
                last_alert_sent_at = completed_at

            if result.get("ok") or is_no_watched_result(result):
                # no watched ships는 감시 목록이 비어 있다는 뜻입니다.
                # 서버/AIS 오류가 아니므로 오류 카운트로 누적하지 않습니다.
                auto_checker_last_success = completed_at
                auto_checker_last_error = "" if result.get("ok") else "no watched ships"
                auto_checker_consecutive_errors = 0
            else:
                auto_checker_last_error = str(result.get("error") or result.get("detail") or "unknown error")
                auto_checker_consecutive_errors += 1
        else:
            auto_checker_last_error = "invalid check result"
            auto_checker_consecutive_errors += 1

    try:
        write_json(SERVER_HEALTH_FILE, build_server_health_payload(include_files=False))
    except Exception:
        pass


def build_server_health_payload(include_files: bool = True) -> Dict[str, Any]:
    watch = read_json(WATCH_FILE, {}) if include_files else {}
    tokens = read_json(TOKENS_FILE, []) if include_files else []
    history = read_json(EVENT_FILE, []) if include_files else []
    state = read_json(AIS_SHIP_STATE_FILE, {}) if include_files else {}
    baseline = read_json(AIS_BASELINE_FILE, {}) if include_files else {}

    current_pid = os.getpid()
    thread_alive = bool(
        auto_checker_thread is not None
        and getattr(auto_checker_thread, "is_alive", lambda: False)()
        and auto_checker_pid == current_pid
    )
    last_completed_age = seconds_since_iso(auto_checker_last_completed or auto_checker_last_run)
    last_success_age = seconds_since_iso(auto_checker_last_success)
    last_alert_age = seconds_since_iso(last_alert_sent_at)

    has_run = bool(auto_checker_last_completed or auto_checker_last_run)
    stale = bool(
        AUTO_CHECK_ENABLED
        and has_run
        and last_completed_age is not None
        and last_completed_age > WATCHDOG_STALE_SECONDS
    )

    summary = watch_summary(watch) if isinstance(watch, dict) else {}
    total_watch_ships = int(summary.get("totalWatchShips") or 0)
    no_watched_idle = total_watch_ships <= 0 or is_no_watched_result(auto_checker_last_result)

    if not AUTO_CHECK_ENABLED:
        status = "disabled"
        status_label = "자동감시 꺼짐"
    elif not thread_alive:
        status = "warning"
        status_label = "자동감시 스레드 확인 필요"
    elif total_watch_ships <= 0:
        status = "idle"
        status_label = "감시 선박 없음 · 앱 재동기화 필요"
    elif not has_run:
        status = "starting"
        status_label = "서버 시작됨 · 첫 AIS 검사 대기 중"
    elif stale:
        status = "stale"
        status_label = "서버 감시 지연"
    elif auto_checker_consecutive_errors > 0 and not no_watched_idle:
        status = "warning"
        status_label = "최근 AIS 검사 오류"
    else:
        status = "ok"
        status_label = "서버 정상 감시 중"


    return {
        "ok": status in ("ok", "starting", "disabled", "idle"),
        "service": "ulsan-ais-fcm-server",
        "version": SERVER_VERSION,
        "status": status,
        "statusLabel": status_label,
        "time": now_iso(),
        "serverStartedAt": SERVER_STARTED_AT,
        "serverUptimeSeconds": seconds_since_iso(SERVER_STARTED_AT),
        "firebaseReady": firebase_ready,
        "firebaseError": firebase_error,
        "autoCheckEnabled": AUTO_CHECK_ENABLED,
        "autoCheckIntervalSeconds": AUTO_CHECK_INTERVAL_SECONDS,
        "watchdogStaleSeconds": WATCHDOG_STALE_SECONDS,
        "autoCheckStarted": auto_checker_started,
        "autoCheckThreadAlive": thread_alive,
        "autoCheckPid": auto_checker_pid,
        "currentPid": current_pid,
        "autoCheckRunCount": auto_checker_run_count,
        "autoCheckLastRun": auto_checker_last_run,
        "autoCheckLastStarted": auto_checker_last_started,
        "autoCheckLastCompleted": auto_checker_last_completed,
        "autoCheckLastCompletedAgeSeconds": last_completed_age,
        "autoCheckLastSuccess": auto_checker_last_success,
        "autoCheckLastSuccessAgeSeconds": last_success_age,
        "autoCheckLastError": auto_checker_last_error,
        "autoCheckConsecutiveErrors": auto_checker_consecutive_errors,
        "lastAlertSentAt": last_alert_sent_at,
        "lastAlertSentAgeSeconds": last_alert_age,
        "lastCheckDurationMs": last_check_duration_ms,
        "lastWatchedCount": last_watched_count,
        "lastAisCount": last_ais_count,
        "lastDetectedCount": last_detected_count,
        "lastEventCount": last_event_count,
        "lastSentCount": last_sent_count,
        "autoCheckLastResult": auto_checker_last_result,
        "tokenCount": safe_len(tokens),
        "historyCount": safe_len(history),
        "shipStateCount": safe_len(state) if isinstance(state, dict) else 0,
        "baselinePendingCount": safe_len(baseline) if isinstance(baseline, dict) else 0,
        **summary,
    }


def init_firebase() -> None:
    global firebase_ready, firebase_error
    if firebase_ready:
        return

    if firebase_admin is None:
        firebase_error = "firebase_admin 패키지를 import하지 못했습니다. requirements.txt를 확인하세요."
        return

    try:
        if firebase_admin._apps:
            firebase_ready = True
            return

        if FIREBASE_SERVICE_ACCOUNT_JSON:
            info = json.loads(FIREBASE_SERVICE_ACCOUNT_JSON)
            cred = credentials.Certificate(info)
        elif FIREBASE_SERVICE_ACCOUNT_FILE:
            cred = credentials.Certificate(FIREBASE_SERVICE_ACCOUNT_FILE)
        else:
            firebase_error = "FIREBASE_SERVICE_ACCOUNT_JSON 환경변수가 없습니다."
            return

        firebase_admin.initialize_app(cred)
        firebase_ready = True
        firebase_error = ""
    except Exception as exc:
        firebase_ready = False
        firebase_error = str(exc)


def send_fcm_to_tokens(target_tokens: List[str], title: str, body: str, data: Dict[str, str] | None = None) -> Dict[str, Any]:
    global last_alert_sent_at
    if not firebase_ready:
        return {
            "ok": False,
            "success": 0,
            "requested": len(target_tokens),
            "errors": [f"firebase not ready: {firebase_error}"],
        }

    success = 0
    errors = []

    for token in target_tokens:
        try:
            msg = messaging.Message(
                token=token,
                notification=messaging.Notification(title=title, body=body),
                data=data or {},
                android=messaging.AndroidConfig(
                    priority="high",
                    notification=messaging.AndroidNotification(sound="default"),
                ),
            )
            message_id = messaging.send(msg)
            print(f"FCM SUCCESS: {message_id}")
            success += 1
            last_alert_sent_at = now_iso()
        except Exception as exc:
            print(f"FCM ERROR: {exc}")
            errors.append(str(exc))

    return {
        "ok": True,
        "requested": len(target_tokens),
        "success": success,
        "errors": errors[:5],
    }


def tokens_for_ship(ship_name: str) -> List[str]:
    normalized = normalize_ship_name(ship_name)
    watch = read_json(WATCH_FILE, {})
    result = []

    for token, item in watch.items():
        ships = unique_ship_list(item.get("ships", [])) if isinstance(item, dict) else []
        notification_off_ships = unique_ship_list(item.get("notificationOffShips", [])) if isinstance(item, dict) else []
        if normalized in ships and normalized not in notification_off_ships:
            result.append(token)

    return result


def notification_disabled_tokens_for_ship(ship_name: str) -> List[str]:
    normalized = normalize_ship_name(ship_name)
    watch = read_json(WATCH_FILE, {})
    result = []

    for token, item in watch.items():
        ships = unique_ship_list(item.get("ships", [])) if isinstance(item, dict) else []
        notification_off_ships = unique_ship_list(item.get("notificationOffShips", [])) if isinstance(item, dict) else []
        if normalized in ships and normalized in notification_off_ships:
            result.append(token)

    return result


def fetch_upa_ais_list() -> List[Dict[str, Any]]:
    url = "https://www.upa.or.kr/abs/cmm/init/getAisListOnTheMap.do"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://www.upa.or.kr/abs/main/mainPage.do",
        "Origin": "https://www.upa.or.kr",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
    }
    data = {
        "bookmarkYn": "N",
        "north": "35.750000",
        "south": "35.150000",
        "east": "129.750000",
        "west": "129.050000",
    }

    response = requests.post(url, headers=headers, data=data, timeout=15)
    response.raise_for_status()
    decoded = response.json()

    if isinstance(decoded, list):
        return [item for item in decoded if isinstance(item, dict)]

    if isinstance(decoded, dict):
        raw_list = (
            decoded.get("vsslList")
            or decoded.get("ships")
            or decoded.get("data")
            or decoded.get("items")
            or decoded.get("vessels")
            or []
        )
        if isinstance(raw_list, list):
            return [item for item in raw_list if isinstance(item, dict)]

    return []


def pick_ship_name(item: Dict[str, Any]) -> str:
    return normalize_ship_name(
        item.get("aisVsslNm")
        or item.get("vsslNm")
        or item.get("shipName")
        or item.get("name")
        or item.get("vslNm")
        or item.get("vesselName")
        or item.get("vsl_eng_nm")
        or ""
    )


def pick_text(item: Dict[str, Any], keys: List[str], default: str = "-") -> str:
    for key in keys:
        value = item.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return default


def pick_float(item: Dict[str, Any], keys: List[str], default: float = 0.0) -> float:
    for key in keys:
        value = item.get(key)
        try:
            if value is not None and str(value).strip():
                return float(str(value).strip())
        except Exception:
            continue
    return default


# 3.1.8 Port-MIS 선박입출항현황 엑셀 업로드/파싱 기능
# 수집 순서:
# Port-MIS 선박입출항현황 → 50000개씩 보기 → 엑셀 다운로드 → /portmis/upload-excel 업로드
# ETA 우선순위 정책:
# 1순위 PORTWISE, 2순위 PORT-MIS. 이 서버 기능은 PORT-MIS ETA를 조기 예보값으로 저장합니다.
PORTMIS_COLUMNS = [
    ("portName", "항명"),
    ("callSign", "호출부호"),
    ("shipName", "선명"),
    ("entryYear", "입항연도"),
    ("entryCount", "입항횟수"),
    ("requestType", "구분"),
    ("inOutPortType", "외항/내항"),
    ("movementType", "입출"),
    ("grossTon", "총톤수"),
    ("arrivalTime", "입항일시"),
    ("departureTime", "출항일시"),
    ("ciqProcessTime", "CIQ수속일자"),
    ("permissionTime", "수리일시"),
    ("voyageType", "항해구분"),
    ("mrn", "MRN 번호"),
    ("berthCode", "계선장소코드"),
    ("berthSubCode", "계선장소세부코드"),
    ("berthName", "계선장소명"),
    ("nextPort", "차항지"),
    ("previousPort", "전출항지"),
    ("shipType", "선박용도"),
    ("koreanCrewCount", "한국인/해기사 선원수"),
    ("foreignCrewCount", "외국인/보통 선원수"),
    ("passengerCount", "승객"),
    ("tugYn", "예선"),
    ("pilotYn", "도선"),
    ("bargeCallSign1", "부선호출부호1"),
    ("bargeCallSign2", "부선호출부호2"),
]


def normalize_ship_match_key(value: Any) -> str:
    """
    선박명 매칭용 키.
    KEOYOUNG SUN 3 / KEOYOUNG SUN3처럼 공백/기호 차이를 줄이기 위해
    영문/숫자/한글만 남기고 대문자로 통일합니다.
    """
    text = normalize_ship_name(value)
    return "".join(ch for ch in text if ch.isalnum() or ("가" <= ch <= "힣"))


def portmis_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def portmis_normalize_datetime(value: Any) -> str:
    text = portmis_cell_text(value)
    if not text:
        return ""
    # Excel이나 WebSquare에서 이미 "2026-06-17 13:00" 형태로 내려옵니다.
    if len(text) >= 16 and text[4:5] == "-" and text[7:8] == "-":
        return text[:16]
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 12:
        return f"{digits[0:4]}-{digits[4:6]}-{digits[6:8]} {digits[8:10]}:{digits[10:12]}"
    if len(digits) == 8:
        return f"{digits[0:4]}-{digits[4:6]}-{digits[6:8]}"
    return text


def portmis_normalize_date(value: Any) -> str:
    text = portmis_cell_text(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 8:
        return f"{digits[0:4]}-{digits[4:6]}-{digits[6:8]}"
    return text


def find_portmis_header_row(ws) -> int:
    """
    Port-MIS 다운로드 엑셀은 보통 12행이 컬럼명이고 13행부터 데이터입니다.
    WebSquare가 만든 xlsx는 dimension 정보가 A1:A1로 잘못 들어오는 경우가 있어
    ws.max_row에 의존하지 않고 앞 40행을 직접 순회합니다.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        values = [portmis_cell_text(value) for value in row]
        joined = "|".join(values)
        if "항명" in joined and "호출부호" in joined and "선명" in joined and "입항일시" in joined:
            return row_idx
    return 12


def extract_portmis_excel_period(ws) -> Dict[str, str]:
    result = {"from": "", "to": "", "printedAt": ""}
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        line = " ".join(portmis_cell_text(v) for v in row if portmis_cell_text(v))
        if "입출항시작일" in line:
            result["from"] = portmis_normalize_date(line)
        elif "입출항종료일" in line:
            result["to"] = portmis_normalize_date(line)
        elif "출력일자" in line:
            result["printedAt"] = portmis_normalize_date(line)
    return result


def parse_portmis_excel_file(file_obj: Any) -> Dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl 패키지가 없습니다. requirements.txt에 openpyxl==3.1.5 를 추가하세요.")

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    # Port-MIS WebSquare 엑셀은 파일 dimension이 A1:A1로 저장되는 경우가 있어
    # reset_dimensions() 후 iter_rows 전체 순회 방식으로 읽어야 2000+행 전체가 잡힙니다.
    try:
        ws.reset_dimensions()
    except Exception:
        pass

    header_row = find_portmis_header_row(ws)
    period = extract_portmis_excel_period(ws)

    items: List[Dict[str, Any]] = []
    seen_record_keys = set()

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        raw_values = [portmis_cell_text(v) for v in row[:len(PORTMIS_COLUMNS)]]
        if not any(raw_values):
            continue

        # 상단/하단 메모나 깨진 행 방지
        port_name = raw_values[0] if len(raw_values) > 0 else ""
        ship_name = raw_values[2] if len(raw_values) > 2 else ""
        if not ship_name or ship_name in ("선명", "선박입출항현황"):
            continue

        item: Dict[str, Any] = {}
        for index, (field, _label) in enumerate(PORTMIS_COLUMNS):
            item[field] = raw_values[index] if index < len(raw_values) else ""

        item["portName"] = portmis_cell_text(item.get("portName"))
        item["callSign"] = portmis_cell_text(item.get("callSign")).upper()
        item["shipName"] = portmis_cell_text(item.get("shipName")).upper()
        item["shipNameRaw"] = portmis_cell_text(raw_values[2] if len(raw_values) > 2 else "")
        item["normalizedShipName"] = normalize_ship_name(item["shipName"])
        item["shipMatchKey"] = normalize_ship_match_key(item["shipName"])
        item["arrivalTime"] = portmis_normalize_datetime(item.get("arrivalTime"))
        item["departureTime"] = portmis_normalize_datetime(item.get("departureTime"))
        item["ciqProcessTime"] = portmis_normalize_datetime(item.get("ciqProcessTime"))
        item["permissionTime"] = portmis_normalize_datetime(item.get("permissionTime"))
        item["source"] = "PORT_MIS_EXCEL"
        item["sourcePriority"] = 2
        item["rowNumber"] = row_idx

        movement = str(item.get("movementType") or "").strip()
        if movement == "입항" and item.get("arrivalTime"):
            item["portmisEta"] = item.get("arrivalTime", "")
            item["eta"] = item.get("arrivalTime", "")
            item["etaSource"] = "PORT_MIS"
            item["etaPriority"] = 2
            item["confidence"] = "PLANNED"
        else:
            item["portmisEta"] = ""
            item["eta"] = ""
            item["etaSource"] = ""
            item["etaPriority"] = 0
            item["confidence"] = "RECORD"

        record_key = "::".join([
            item.get("portName", ""),
            item.get("callSign", ""),
            item.get("shipMatchKey", ""),
            item.get("entryYear", ""),
            item.get("entryCount", ""),
            item.get("movementType", ""),
            item.get("arrivalTime", ""),
            item.get("departureTime", ""),
        ])
        if record_key in seen_record_keys:
            continue
        seen_record_keys.add(record_key)
        items.append(item)

    wb.close()

    uploaded_at = now_iso()
    port_counts: Dict[str, int] = {}
    movement_counts: Dict[str, int] = {}
    for item in items:
        port = str(item.get("portName") or "미상")
        movement = str(item.get("movementType") or "미상")
        port_counts[port] = port_counts.get(port, 0) + 1
        movement_counts[movement] = movement_counts.get(movement, 0) + 1

    return {
        "ok": True,
        "source": "PORT_MIS_EXCEL",
        "version": SERVER_VERSION,
        "uploadedAt": uploaded_at,
        "from": period.get("from", ""),
        "to": period.get("to", ""),
        "printedAt": period.get("printedAt", ""),
        "sheetName": ws.title,
        "headerRow": header_row,
        "count": len(items),
        "portCounts": dict(sorted(port_counts.items(), key=lambda kv: kv[0])),
        "movementCounts": dict(sorted(movement_counts.items(), key=lambda kv: kv[0])),
        "etaPolicy": {
            "representativeEtaPriority": ["PORTWISE", "PORT_MIS"],
            "portmisPriority": 2,
            "description": "PORT-MIS는 1~2주 전 조기 입항예정/선석회의 기반 데이터로 사용하고, PORTWISE ETA가 있으면 대표 ETA는 PORTWISE로 교체합니다.",
        },
        "items": items,
    }


# 3.1.9 Port-MIS auto-download xlsx header parser.
# Latest auto collector files have header row at row 12 and 41 columns from A to AO.
PORTMIS_HEADER_SPECS = [
    {"field": "portName", "position": 1, "labels": ["항명"]},
    {"field": "callSign", "position": 2, "labels": ["호출부호"]},
    {"field": "shipName", "position": 3, "labels": ["선명"]},
    {"field": "entryYear", "position": 4, "labels": ["입항년도", "입항횟수"]},
    {"field": "entryCount", "position": 5, "labels": ["입항횟수"]},
    {"field": "requestType", "position": 6, "labels": ["구분"]},
    {"field": "inOutPortType", "position": 7, "labels": ["외내"]},
    {"field": "movementType", "position": 8, "labels": ["입출"]},
    {"field": "grossTon", "position": 9, "labels": ["총톤수"]},
    {"field": "internationalTon", "position": 10, "labels": ["국제톤수"]},
    {"field": "billingTon", "position": 11, "labels": ["징수톤수"]},
    {"field": "arrivalTime", "position": 12, "labels": ["입항일시"]},
    {"field": "departureTime", "position": 13, "labels": ["출항일시"]},
    {"field": "ciqProcessTime", "position": 14, "labels": ["CIQ수속일자"]},
    {"field": "permissionTime", "position": 15, "labels": ["수리일시"]},
    {"field": "purpose", "position": 16, "labels": ["입항목적"]},
    {"field": "voyageType", "position": 17, "labels": ["항해구분"]},
    {"field": "mrn", "position": 18, "labels": ["MRN번호", "MRN"]},
    {"field": "cargoType", "position": 19, "labels": ["적재화물"]},
    {"field": "cargoTon", "position": 20, "labels": ["적재톤수"]},
    {"field": "nationalityCode", "position": 21, "labels": ["국적코드"]},
    {"field": "nationalityName", "position": 22, "labels": ["국적명"]},
    {"field": "berthCode", "position": 23, "labels": ["계선장소코드"]},
    {"field": "berthSubCode", "position": 24, "labels": ["계선장소번호"]},
    {"field": "berthName", "position": 25, "labels": ["계선장소명"]},
    {"field": "discountRate", "position": 26, "labels": ["할인율"]},
    {"field": "discountReason", "position": 27, "labels": ["할인사유"]},
    {"field": "nextPort", "position": 28, "labels": ["차항지"]},
    {"field": "previousPort", "position": 29, "labels": ["전출항지"]},
    {"field": "shippingCompanyCode", "position": 30, "labels": ["선사코드"]},
    {"field": "shippingCompanyName", "position": 31, "labels": ["선사명"]},
    {"field": "agentCode", "position": 32, "labels": ["대리점코드"]},
    {"field": "agentName", "position": 33, "labels": ["대리점명"]},
    {"field": "shipType", "position": 34, "labels": ["선박용도"]},
    {"field": "koreanCrewCount", "position": 35, "labels": ["한국인선원수", "해기사선원수"]},
    {"field": "foreignCrewCount", "position": 36, "labels": ["외국인선원수", "보통선원수"]},
    {"field": "passengerCount", "position": 37, "labels": ["승객"]},
    {"field": "tugYn", "position": 38, "labels": ["예선"]},
    {"field": "pilotYn", "position": 39, "labels": ["도선"]},
    {"field": "bargeCallSign1", "position": 40, "labels": ["부선호출부호1"]},
    {"field": "bargeCallSign2", "position": 41, "labels": ["부선호출부호2"]},
]


def portmis_header_key(value: Any) -> str:
    text = portmis_cell_text(value)
    for old, new in (
        ("\n", ""),
        ("\r", ""),
        ("\t", ""),
        (" ", ""),
        ("/", ""),
        (":", ""),
        ("(", ""),
        (")", ""),
        ("<br/>", ""),
        ("<br>", ""),
    ):
        text = text.replace(old, new)
    return text.upper()


def portmis_header_matches(header_value: Any, labels: List[str]) -> bool:
    header = portmis_header_key(header_value)
    if not header:
        return False
    for label in labels:
        key = portmis_header_key(label)
        if key and (key in header or header in key):
            return True
    return False


def portmis_header_row_values(ws, row_idx: int) -> List[str]:
    row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), [])
    return [portmis_cell_text(value) for value in row]


def portmis_row_looks_like_header(values: List[str]) -> bool:
    joined = "|".join(portmis_header_key(value) for value in values)
    return all(token in joined for token in ("항명", "호출부호", "선명", "입출"))


def find_portmis_header_row(ws) -> int:
    row12 = portmis_header_row_values(ws, 12)
    if portmis_row_looks_like_header(row12):
        return 12

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        values = [portmis_cell_text(value) for value in row]
        if portmis_row_looks_like_header(values):
            return row_idx
    return 12


def build_portmis_column_indexes(headers: List[str]) -> Dict[str, int]:
    indexes: Dict[str, int] = {}
    used = set()

    for spec in PORTMIS_HEADER_SPECS:
        field = str(spec["field"])
        labels = list(spec["labels"])
        preferred = int(spec["position"]) - 1

        if preferred < len(headers) and portmis_header_matches(headers[preferred], labels):
            indexes[field] = preferred
            used.add(preferred)
            continue

        found = None
        for idx, header in enumerate(headers):
            if idx in used:
                continue
            if portmis_header_matches(header, labels):
                found = idx
                break

        if found is None:
            found = preferred

        indexes[field] = found
        used.add(found)

    return indexes


def portmis_get_by_index(values: List[str], index: int) -> str:
    if index < 0 or index >= len(values):
        return ""
    return portmis_cell_text(values[index])


def normalize_ship_match_key(value: Any) -> str:
    text = normalize_ship_name(value)
    return "".join(ch for ch in text if ch.isalnum())


def portmis_date_part(value: Any) -> str:
    text = portmis_normalize_datetime(value)
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        return text[:10]
    return ""


def extract_portmis_excel_period(ws) -> Dict[str, str]:
    result = {"from": "", "to": "", "printedAt": ""}
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        line = " ".join(portmis_cell_text(v) for v in row if portmis_cell_text(v))
        if not line:
            continue
        normalized = portmis_header_key(line)
        if "시작" in normalized or "FROM" in normalized:
            result["from"] = portmis_normalize_date(line)
        elif "종료" in normalized or "TO" in normalized:
            result["to"] = portmis_normalize_date(line)
        elif "출력" in normalized or "PRINT" in normalized:
            result["printedAt"] = portmis_normalize_date(line)
    return result


def parse_portmis_excel_file(file_obj: Any) -> Dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl 패키지가 없습니다. requirements.txt에 openpyxl==3.1.5 를 추가하세요.")

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    try:
        ws.reset_dimensions()
    except Exception:
        pass

    header_row = find_portmis_header_row(ws)
    headers = portmis_header_row_values(ws, header_row)
    column_indexes = build_portmis_column_indexes(headers)
    period = extract_portmis_excel_period(ws)

    items: List[Dict[str, Any]] = []
    seen_record_keys = set()
    blank_streak = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        raw_values = [portmis_cell_text(v) for v in row]
        if not any(raw_values):
            blank_streak += 1
            if blank_streak >= 50 and items:
                break
            continue
        blank_streak = 0

        item: Dict[str, Any] = {}
        for spec in PORTMIS_HEADER_SPECS:
            field = str(spec["field"])
            item[field] = portmis_get_by_index(raw_values, column_indexes.get(field, -1))

        ship_name_raw = portmis_cell_text(item.get("shipName"))
        port_name = portmis_cell_text(item.get("portName"))
        if not ship_name_raw or portmis_header_matches(ship_name_raw, ["선명"]):
            continue
        if not port_name or portmis_header_matches(port_name, ["항명"]):
            continue

        item["portName"] = port_name
        item["callSign"] = portmis_cell_text(item.get("callSign")).upper()
        item["shipNameRaw"] = ship_name_raw
        item["shipName"] = ship_name_raw.upper()
        item["normalizedShipName"] = normalize_ship_name(item["shipName"])
        item["shipMatchKey"] = normalize_ship_match_key(item["shipName"])

        item["arrivalTime"] = portmis_normalize_datetime(item.get("arrivalTime"))
        item["departureTime"] = portmis_normalize_datetime(item.get("departureTime"))
        item["ciqProcessTime"] = portmis_normalize_datetime(item.get("ciqProcessTime"))
        item["permissionTime"] = portmis_normalize_datetime(item.get("permissionTime"))

        item["source"] = "PORT_MIS_EXCEL"
        item["sourcePriority"] = 2
        item["rowNumber"] = row_idx

        movement = str(item.get("movementType") or "").strip()
        if "입항" in movement and item.get("arrivalTime"):
            item["portmisEta"] = item.get("arrivalTime", "")
            item["eta"] = item.get("arrivalTime", "")
            item["etaSource"] = "PORT_MIS"
            item["etaPriority"] = 2
            item["confidence"] = "PLANNED"
        elif "출항" in movement:
            item["portmisEta"] = ""
            item["eta"] = ""
            item["etaSource"] = ""
            item["etaPriority"] = 0
            item["confidence"] = "DEPARTURE_RECORD"
        else:
            item["portmisEta"] = ""
            item["eta"] = ""
            item["etaSource"] = ""
            item["etaPriority"] = 0
            item["confidence"] = "RECORD"

        record_key = "::".join([
            item.get("portName", ""),
            item.get("callSign", ""),
            item.get("shipMatchKey", ""),
            item.get("entryYear", ""),
            item.get("entryCount", ""),
            item.get("movementType", ""),
            item.get("arrivalTime", ""),
            item.get("departureTime", ""),
        ])
        if record_key in seen_record_keys:
            continue
        seen_record_keys.add(record_key)
        items.append(item)

    sheet_name = ws.title
    wb.close()

    date_parts = []
    for item in items:
        movement = str(item.get("movementType") or "").strip()
        fields = ["arrivalTime"] if "입항" in movement else ["departureTime"] if "출항" in movement else ["arrivalTime", "departureTime"]
        for field in fields:
            part = portmis_date_part(item.get(field))
            if part:
                date_parts.append(part)
    if date_parts:
        if not period.get("from"):
            period["from"] = min(date_parts)
        if not period.get("to"):
            period["to"] = max(date_parts)

    uploaded_at = now_iso()
    port_counts: Dict[str, int] = {}
    movement_counts: Dict[str, int] = {}
    for item in items:
        port = str(item.get("portName") or "미상")
        movement = str(item.get("movementType") or "미상")
        port_counts[port] = port_counts.get(port, 0) + 1
        movement_counts[movement] = movement_counts.get(movement, 0) + 1

    return {
        "ok": True,
        "source": "PORT_MIS_EXCEL",
        "version": SERVER_VERSION,
        "uploadedAt": uploaded_at,
        "from": period.get("from", ""),
        "to": period.get("to", ""),
        "printedAt": period.get("printedAt", ""),
        "sheetName": sheet_name,
        "headerRow": header_row,
        "columnCount": max((int(spec["position"]) for spec in PORTMIS_HEADER_SPECS), default=0),
        "columnMap": {field: index + 1 for field, index in column_indexes.items()},
        "count": len(items),
        "portCounts": dict(sorted(port_counts.items(), key=lambda kv: kv[0])),
        "movementCounts": dict(sorted(movement_counts.items(), key=lambda kv: kv[0])),
        "sample": items[:3],
        "etaPolicy": {
            "representativeEtaPriority": ["PORTWISE", "PORT_MIS"],
            "portmisPriority": 2,
            "description": "PORT-MIS는 사전 예정 자료이며, PORTWISE ETA가 있으면 앱에서 PORTWISE를 대표 ETA로 우선 표시합니다.",
        },
        "items": items,
    }


def portmis_status_payload(include_items: bool = False) -> Dict[str, Any]:
    data = read_json(PORTMIS_FILE, {})
    if not isinstance(data, dict) or not data:
        return {
            "ok": True,
            "available": False,
            "source": "PORT_MIS_EXCEL",
            "count": 0,
            "message": "아직 업로드된 Port-MIS 엑셀 데이터가 없습니다.",
            "time": now_iso(),
        }

    payload = {
        "ok": True,
        "available": True,
        "source": data.get("source", "PORT_MIS_EXCEL"),
        "uploadedAt": data.get("uploadedAt", ""),
        "from": data.get("from", ""),
        "to": data.get("to", ""),
        "printedAt": data.get("printedAt", ""),
        "count": int(data.get("count") or len(data.get("items", []) if isinstance(data.get("items"), list) else [])),
        "portCounts": data.get("portCounts", {}),
        "movementCounts": data.get("movementCounts", {}),
        "etaPolicy": data.get("etaPolicy", {}),
        "time": now_iso(),
    }
    if include_items:
        payload["items"] = data.get("items", []) if isinstance(data.get("items"), list) else []
    return payload


def portmis_filtered_items(args: Dict[str, Any]) -> Dict[str, Any]:
    data = read_json(PORTMIS_FILE, {})
    if not isinstance(data, dict) or not isinstance(data.get("items"), list):
        return {
            "ok": True,
            "available": False,
            "count": 0,
            "items": [],
            "message": "아직 업로드된 Port-MIS 엑셀 데이터가 없습니다.",
            "time": now_iso(),
        }

    items = [item for item in data.get("items", []) if isinstance(item, dict)]
    original_count = len(items)

    port = str(args.get("port") or "").strip()
    if port:
        items = [item for item in items if port in str(item.get("portName") or "")]

    movement = str(args.get("movement") or args.get("movementType") or "").strip()
    if movement:
        items = [item for item in items if movement == str(item.get("movementType") or "").strip()]

    ship_query = str(args.get("ship") or args.get("shipName") or "").strip()
    if ship_query:
        q_key = normalize_ship_match_key(ship_query)
        items = [
            item for item in items
            if q_key in str(item.get("shipMatchKey") or "") or ship_query.upper() in str(item.get("shipName") or "")
        ]

    tracked_only = str(args.get("trackedOnly") or args.get("tracked") or "").strip().lower() in ("1", "true", "yes", "y", "on")
    token = str(args.get("token") or "").strip()
    tracked_keys = set()
    if tracked_only:
        watch = read_json(WATCH_FILE, {})
        if isinstance(watch, dict):
            if token and isinstance(watch.get(token), dict):
                tracked_keys.update(normalize_ship_match_key(name) for name in unique_ship_list(watch[token].get("ships", [])))
            else:
                for _device_token, entry in watch.items():
                    if isinstance(entry, dict):
                        tracked_keys.update(normalize_ship_match_key(name) for name in unique_ship_list(entry.get("ships", [])))
        tracked_keys.discard("")
        items = [item for item in items if str(item.get("shipMatchKey") or "") in tracked_keys]

    eta_only = str(args.get("etaOnly") or "").strip().lower() in ("1", "true", "yes", "y", "on")
    if eta_only:
        items = [item for item in items if str(item.get("movementType") or "").strip() == "입항" and item.get("arrivalTime")]

    try:
        limit = int(args.get("limit") or 0)
    except Exception:
        limit = 0
    if limit > 0:
        items = items[:limit]

    return {
        "ok": True,
        "available": True,
        "source": data.get("source", "PORT_MIS_EXCEL"),
        "uploadedAt": data.get("uploadedAt", ""),
        "from": data.get("from", ""),
        "to": data.get("to", ""),
        "printedAt": data.get("printedAt", ""),
        "originalCount": original_count,
        "count": len(items),
        "filters": {
            "port": port,
            "movement": movement,
            "ship": ship_query,
            "trackedOnly": tracked_only,
            "tokenScoped": bool(token),
            "etaOnly": eta_only,
            "limit": limit,
        },
        "etaPolicy": data.get("etaPolicy", {}),
        "items": items,
        "time": now_iso(),
    }



@app.before_request
def before_request() -> None:
    init_firebase()
    ensure_auto_checker_running()


@app.get("/")
def index():
    return jsonify(build_server_health_payload())


@app.get("/health")
def health():
    # 외부 모니터링/Render 헬스체크용. API 키 없이도 최소 상태를 확인할 수 있게 둡니다.
    payload = build_server_health_payload()
    return jsonify({
        "ok": payload.get("ok", True),
        "service": payload.get("service"),
        "version": payload.get("version"),
        "status": payload.get("status"),
        "statusLabel": payload.get("statusLabel"),
        "firebaseReady": payload.get("firebaseReady"),
        "autoCheckThreadAlive": payload.get("autoCheckThreadAlive"),
        "autoCheckLastCompleted": payload.get("autoCheckLastCompleted"),
        "autoCheckLastCompletedAgeSeconds": payload.get("autoCheckLastCompletedAgeSeconds"),
        "autoCheckRunCount": payload.get("autoCheckRunCount"),
        "time": payload.get("time"),
    })


@app.get("/server-health")
def server_health():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    ensure_auto_checker_running()
    return jsonify(build_server_health_payload())


@app.post("/watchdog-check")
def watchdog_check():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    ensure_auto_checker_running()
    payload = request.get_json(silent=True) or {}
    force = bool(payload.get("force", False))
    started = time.time()
    result = perform_ais_check_once(force=force, source="watchdog")
    update_health_from_check_result(result, int((time.time() - started) * 1000))
    return jsonify({
        "ok": bool(result.get("ok")),
        "result": result,
        "health": build_server_health_payload(),
        "time": now_iso(),
    }), 200 if result.get("ok") else 500


@app.post("/register-token")
def register_token():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    token = str(payload.get("token", "")).strip()
    if not token:
        return jsonify({"ok": False, "error": "token is required"}), 400

    tokens: List[Dict[str, Any]] = read_json(TOKENS_FILE, [])
    now = now_iso()
    found = False

    for item in tokens:
        if item.get("token") == token:
            item.update({
                "app": payload.get("app", item.get("app", "ulsan_ais_mobile")),
                "platform": payload.get("platform", item.get("platform", "android")),
                "device": payload.get("device", item.get("device", "android")),
                "version": payload.get("version", item.get("version", "")),
                "lastSeenAt": now,
            })
            found = True
            break

    if not found:
        tokens.append({
            "token": token,
            "app": payload.get("app", "ulsan_ais_mobile"),
            "platform": payload.get("platform", "android"),
            "device": payload.get("device", "android"),
            "version": payload.get("version", ""),
            "createdAt": now,
            "lastSeenAt": now,
        })

    write_json(TOKENS_FILE, tokens)
    return jsonify({"ok": True, "registered": True, "tokenCount": len(tokens), "time": now})


@app.get("/tokens/count")
def token_count():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401
    tokens = read_json(TOKENS_FILE, [])
    return jsonify({"ok": True, "tokenCount": len(tokens)})


@app.post("/send-test")
def send_test():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401
    if not firebase_ready:
        return jsonify({"ok": False, "error": "firebase not ready", "firebaseError": firebase_error}), 500

    payload = request.get_json(silent=True) or {}
    title = str(payload.get("title", "🚢 울산항 AIS 테스트 알림"))
    body = str(payload.get("body", "FCM 서버에서 보낸 테스트 푸시입니다."))
    token = str(payload.get("token", "")).strip()

    if token:
        target_tokens = [token]
    else:
        stored = read_json(TOKENS_FILE, [])
        target_tokens = [
            str(item.get("token", "")).strip()
            for item in stored
            if str(item.get("token", "")).strip()
        ]

    if not target_tokens:
        return jsonify({"ok": False, "error": "no tokens registered"}), 400

    result = send_fcm_to_tokens(
        target_tokens,
        title,
        body,
        {
            "source": "ulsan_ais_fcm_server",
            "eventType": "test",
            "sentAt": now_iso(),
        },
    )

    history = read_json(EVENT_FILE, [])
    history.insert(0, {
        "type": "test",
        "title": title,
        "body": body,
        "success": result["success"],
        "errors": result["errors"],
        "time": now_iso(),
    })
    write_json(EVENT_FILE, history[:200])

    return jsonify(result)


@app.post("/test-ship-alert")
def test_ship_alert():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401
    if not firebase_ready:
        return jsonify({"ok": False, "error": "firebase not ready", "firebaseError": firebase_error}), 500

    payload = request.get_json(silent=True) or {}
    ship_name = normalize_ship_name(payload.get("shipName", ""))
    title = str(payload.get("title", f"🚢 {ship_name} 감시 알림"))
    body = str(payload.get("body", f"{ship_name} 테스트 선박 이벤트가 발생했습니다."))

    if not ship_name:
        return jsonify({"ok": False, "error": "shipName is required"}), 400

    target_tokens = tokens_for_ship(ship_name)

    if not target_tokens:
        return jsonify({
            "ok": False,
            "error": "no watchers for this ship",
            "shipName": ship_name,
            "targetCount": 0,
        }), 404

    result = send_fcm_to_tokens(
        target_tokens,
        title,
        body,
        {
            "source": "ulsan_ais_fcm_server",
            "eventType": "ship_test_alert",
            "shipName": ship_name,
            "sentAt": now_iso(),
        },
    )

    print(f"SHIP ALERT result ship={ship_name} result={result}")

    history = read_json(EVENT_FILE, [])
    history.insert(0, {
        "type": "ship_test_alert",
        "shipName": ship_name,
        "title": title,
        "body": body,
        "targetCount": len(target_tokens),
        "success": result["success"],
        "errors": result["errors"],
        "time": now_iso(),
    })
    write_json(EVENT_FILE, history[:200])

    return jsonify({
        "ok": True,
        "shipName": ship_name,
        "targetCount": len(target_tokens),
        "success": result["success"],
        "errors": result["errors"],
        "time": now_iso(),
    })




def is_underway_like(status: str, speed: float) -> bool:
    value = str(status).upper()
    return (
        speed >= 3.0
        or "항해" in value
        or "UNDER" in value
        or "SAILING" in value
        or "NAVIGATION" in value
    )


def is_stopped_like(status: str, speed: float) -> bool:
    value = str(status).upper()
    return (
        speed <= 1.0
        or "정박" in value
        or "접안" in value
        or "묘박" in value
        or "투묘" in value
        or "MOORED" in value
        or "ANCHOR" in value
        or "BERTH" in value
        or "DOCK" in value
    )


# 3.0.5 M/E 묘박지 좌표 기반 판정
# 모바일 지도에 표시한 실제 M1~M7, E1~E3 polygon 좌표를 서버에도 동일하게 넣었습니다.
# AIS 좌표가 이 polygon 안에 들어오면 "M3 묘박지"처럼 세부 구역명으로 표시됩니다.
ANCHORAGE_AREAS = [
    {
        "name": "M1",
        "coords": [
            (35.500222, 129.394167),
            (35.495583, 129.394778),
            (35.495583, 129.404611),
            (35.498556, 129.402194),
        ],
    },
    {
        "name": "M2",
        "coords": [
            (35.495583, 129.394778),
            (35.492694, 129.395139),
            (35.491861, 129.395278),
            (35.491861, 129.400361),
            (35.495583, 129.399889),
        ],
    },
    {
        "name": "M3",
        "coords": [
            (35.491861, 129.405722),
            (35.493167, 129.406556),
            (35.495583, 129.404611),
            (35.495583, 129.399889),
            (35.491861, 129.400361),
        ],
    },
    {
        "name": "M4",
        "coords": [
            (35.491861, 129.395278),
            (35.488056, 129.395917),
            (35.488056, 129.403333),
            (35.491861, 129.405722),
        ],
    },
    {
        "name": "M5",
        "coords": [
            (35.488056, 129.395917),
            (35.484250, 129.396583),
            (35.484250, 129.402639),
            (35.486889, 129.402611),
            (35.488056, 129.403333),
        ],
    },
    {
        "name": "M6",
        "coords": [
            (35.484250, 129.396583),
            (35.480444, 129.397222),
            (35.480444, 129.402639),
            (35.484250, 129.402639),
        ],
    },
    {
        "name": "M7",
        "coords": [
            (35.480444, 129.397222),
            (35.476417, 129.397917),
            (35.476417, 129.402639),
            (35.480444, 129.402639),
        ],
    },
    {
        "name": "E1",
        "coords": [
            (35.466389, 129.414278), (35.466389, 129.426306), (35.466452, 129.427450),
            (35.466490, 129.428587), (35.466502, 129.429717), (35.466489, 129.430839),
            (35.466451, 129.431954), (35.466388, 129.433060), (35.466302, 129.434159),
            (35.466192, 129.435249), (35.466058, 129.436330), (35.465901, 129.437403),
            (35.465721, 129.438467), (35.465519, 129.439522), (35.465295, 129.440567),
            (35.465049, 129.441603), (35.464781, 129.442629), (35.464492, 129.443645),
            (35.464183, 129.444651), (35.463853, 129.445647), (35.463503, 129.446632),
            (35.463133, 129.447607), (35.462744, 129.448570), (35.462336, 129.449523),
            (35.461909, 129.450464), (35.461463, 129.451394), (35.461000, 129.452312),
            (35.460518, 129.453218), (35.460020, 129.454112), (35.459504, 129.454994),
            (35.458971, 129.455863), (35.458422, 129.456720), (35.457857, 129.457564),
            (35.457276, 129.458395), (35.456680, 129.459212), (35.456069, 129.460016),
            (35.455443, 129.460807), (35.454803, 129.461583), (35.454148, 129.462346),
            (35.453480, 129.463094), (35.452799, 129.463828), (35.452104, 129.464547),
            (35.451397, 129.465252), (35.437111, 129.410972), (35.462056, 129.401306),
        ],
    },
    {
        "name": "E2",
        "coords": [
            (35.437111, 129.410972), (35.451397, 129.465252), (35.450677, 129.465941),
            (35.449945, 129.466615), (35.449202, 129.467274), (35.448447, 129.467917),
            (35.447681, 129.468545), (35.446905, 129.469156), (35.446118, 129.469751),
            (35.445321, 129.470330), (35.444514, 129.470892), (35.443698, 129.471438),
            (35.442873, 129.471966), (35.442040, 129.472477), (35.441198, 129.472971),
            (35.440348, 129.473448), (35.439490, 129.473906), (35.438625, 129.474347),
            (35.437753, 129.474769), (35.436875, 129.475173), (35.435990, 129.475559),
            (35.435099, 129.475926), (35.434202, 129.476274), (35.433300, 129.476602),
            (35.432393, 129.476912), (35.431482, 129.477202), (35.430566, 129.477472),
            (35.429646, 129.477722), (35.428722, 129.477952), (35.427796, 129.478162),
            (35.426866, 129.478351), (35.425934, 129.478520), (35.424999, 129.478668),
            (35.424062, 129.478794), (35.423124, 129.478899), (35.420194, 129.417528),
        ],
    },
    {
        "name": "E3",
        "coords": [
            (35.420194, 129.417528), (35.423124, 129.478899), (35.422185, 129.478983),
            (35.421244, 129.479045), (35.420303, 129.479086), (35.419362, 129.479104),
            (35.418420, 129.479100), (35.417479, 129.479073), (35.416539, 129.479024),
            (35.415600, 129.478952), (35.414662, 129.478857), (35.413727, 129.478738),
            (35.412793, 129.478596), (35.411861, 129.478431), (35.410932, 129.478241),
            (35.410007, 129.478028), (35.409085, 129.477790), (35.408166, 129.477528),
            (35.407252, 129.477242), (35.406342, 129.476930), (35.405437, 129.476593),
            (35.404537, 129.476232), (35.403642, 129.475845), (35.402753, 129.475432),
            (35.401871, 129.474993), (35.400994, 129.474529), (35.400125, 129.474038),
            (35.399262, 129.473521), (35.398408, 129.472977), (35.397560, 129.472406),
            (35.396721, 129.471809), (35.395891, 129.471184), (35.395069, 129.470532),
            (35.394257, 129.469853), (35.393453, 129.469145), (35.392660, 129.468410),
            (35.391877, 129.467646), (35.391104, 129.466854), (35.390342, 129.466034),
            (35.389592, 129.465185), (35.388852, 129.464307), (35.388125, 129.463400),
            (35.387410, 129.462463), (35.386707, 129.461497), (35.386017, 129.460501),
            (35.385340, 129.459475), (35.384677, 129.458420), (35.384028, 129.457333),
            (35.403056, 129.424167),
        ],
    },
]


def point_in_polygon(lat: float, lon: float, coords: List[tuple[float, float]]) -> bool:
    """
    Ray-casting 방식의 polygon 내부 판정.
    coords는 [(lat, lon), ...]이고, 계산에서는 lon을 x, lat을 y로 사용합니다.
    """
    if not coords:
        return False

    x = lon
    y = lat
    inside = False
    j = len(coords) - 1

    for i in range(len(coords)):
        yi, xi = coords[i]
        yj, xj = coords[j]
        intersects = ((yi > y) != (yj > y)) and (
            x < (xj - xi) * (y - yi) / ((yj - yi) if (yj - yi) != 0 else 1e-12) + xi
        )
        if intersects:
            inside = not inside
        j = i

    return inside


def _distance_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """작은 거리 비교용 Haversine 거리(km)."""
    radius_km = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * radius_km * math.atan2(math.sqrt(a), math.sqrt(max(0.0, 1 - a)))


def _polygon_center(coords: List[tuple[float, float]]) -> tuple[float, float]:
    if not coords:
        return 0.0, 0.0
    return sum(p[0] for p in coords) / len(coords), sum(p[1] for p in coords) / len(coords)


def anchorage_zone_from_lat_lon(lat: float, lon: float) -> str | None:
    if lat == 0 or lon == 0:
        return None

    # 1순위: 실제 M/E polygon 내부 판정
    for area in ANCHORAGE_AREAS:
        if point_in_polygon(lat, lon, area["coords"]):
            return f'{area["name"]} 묘박지'

    # 2순위: AIS 좌표 흔들림 보정. M구역은 작으므로 250m 이내만 "인근"으로 인정합니다.
    # E구역은 면적이 넓고 서로 인접하므로 polygon 밖은 보수적으로 인근 판정을 하지 않습니다.
    nearest_name = ""
    nearest_km = 999.0
    for area in ANCHORAGE_AREAS:
        name = str(area["name"])
        if not name.startswith("M"):
            continue
        center_lat, center_lon = _polygon_center(area["coords"])
        distance = _distance_km(lat, lon, center_lat, center_lon)
        if distance < nearest_km:
            nearest_km = distance
            nearest_name = name

    if nearest_name and nearest_km <= 0.25:
        return f"{nearest_name} 묘박지 인근"

    return None


def anchorage_debug_payload(lat: float, lon: float) -> Dict[str, Any]:
    zone = anchorage_zone_from_lat_lon(lat, lon)
    distances = []
    for area in ANCHORAGE_AREAS:
        center_lat, center_lon = _polygon_center(area["coords"])
        distances.append({
            "name": area["name"],
            "centerLat": round(center_lat, 6),
            "centerLon": round(center_lon, 6),
            "distanceKm": round(_distance_km(lat, lon, center_lat, center_lon), 3),
            "inside": point_in_polygon(lat, lon, area["coords"]),
        })
    distances.sort(key=lambda item: item["distanceKm"])
    return {
        "zone": zone,
        "nearest": distances[:3],
    }

# 3.0.6 SK / S-OIL 부이 좌표 기반 판정
# 2-2순위 작업 범위입니다.
# AIS 좌표가 아래 반경 안에 들어오면 "SK_B#2 부이"처럼 세부 구역명으로 표시됩니다.
# 반경은 AIS 위치 흔들림과 지도 표시 오차를 감안해 실제 표식보다 조금 여유 있게 잡았습니다.
BUOY_AREAS = [
    {"name": "SK_B#2", "display": "SK_B#2 부이", "lat": 35.438800, "lon": 129.393400, "radius_m": 350},
    {"name": "SK_B#3", "display": "SK_B#3 부이", "lat": 35.429500, "lon": 129.393300, "radius_m": 350},
    {"name": "S.OIL_B#1", "display": "S.OIL_B#1 부이", "lat": 35.407100, "lon": 129.395400, "radius_m": 400},
    {"name": "S.OIL_B#2", "display": "S.OIL_B#2 부이", "lat": 35.396700, "lon": 129.393100, "radius_m": 400},
]


def buoy_zone_from_lat_lon(lat: float, lon: float) -> str | None:
    if lat == 0 or lon == 0:
        return None

    nearest = None
    nearest_km = 999.0

    for area in BUOY_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        radius_km = float(area["radius_m"]) / 1000.0
        if distance_km <= radius_km and distance_km < nearest_km:
            nearest = area
            nearest_km = distance_km

    if nearest is None:
        return None

    return str(nearest["display"])


def buoy_debug_payload(lat: float, lon: float) -> Dict[str, Any]:
    zone = buoy_zone_from_lat_lon(lat, lon)
    distances = []
    for area in BUOY_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        distances.append({
            "name": area["name"],
            "display": area["display"],
            "lat": area["lat"],
            "lon": area["lon"],
            "radiusM": area["radius_m"],
            "distanceKm": round(distance_km, 3),
            "inside": distance_km <= float(area["radius_m"]) / 1000.0,
        })
    distances.sort(key=lambda item: item["distanceKm"])
    return {
        "zone": zone,
        "nearest": distances[:4],
    }


# 3.0.7 울산 본항 주요 부두 좌표 기반 판정
# 2-3순위 작업 범위입니다.
# AIS 좌표가 아래 부두 기준점 반경 안에 들어오면 "울산본항 5부두"처럼 세부 구역명으로 표시됩니다.
# 실제 AIS 위치는 선박 중앙/안테나 위치로 잡히기 때문에 부두 표식점보다 약간 넓은 반경을 사용합니다.
MAIN_BERTH_AREAS = [
    {"name": "울산본항 1부두", "lat": 35.531730, "lon": 129.373200, "radius_m": 520},
    {"name": "울산본항 2부두", "lat": 35.526781, "lon": 129.372631, "radius_m": 500},
    {"name": "울산본항 3부두", "lat": 35.522157, "lon": 129.374700, "radius_m": 480},
    {"name": "울산본항 4부두", "lat": 35.520900, "lon": 129.375300, "radius_m": 460},
    {"name": "울산본항 5부두", "lat": 35.518890, "lon": 129.373743, "radius_m": 480},
    {"name": "울산본항 6부두", "lat": 35.517085, "lon": 129.378600, "radius_m": 480},
    {"name": "울산본항 7부두", "lat": 35.516580, "lon": 129.382712, "radius_m": 480},
    {"name": "울산본항 8부두", "lat": 35.515100, "lon": 129.385400, "radius_m": 480},
    {"name": "울산본항 9부두", "lat": 35.511400, "lon": 129.385700, "radius_m": 520},

    {"name": "울산본항 SK1부두", "lat": 35.502700, "lon": 129.363200, "radius_m": 520},
    {"name": "울산본항 SK2부두", "lat": 35.498600, "lon": 129.365800, "radius_m": 520},
    {"name": "울산본항 SK3부두", "lat": 35.494600, "lon": 129.382900, "radius_m": 520},
    {"name": "울산본항 SK4부두", "lat": 35.492700, "lon": 129.384200, "radius_m": 520},
    {"name": "울산본항 SK5부두", "lat": 35.489200, "lon": 129.386300, "radius_m": 520},
    {"name": "울산본항 SK6부두", "lat": 35.485500, "lon": 129.390700, "radius_m": 520},
    {"name": "울산본항 SK7부두", "lat": 35.482400, "lon": 129.390700, "radius_m": 520},
    {"name": "울산본항 SK8부두", "lat": 35.479100, "lon": 129.390600, "radius_m": 520},

    {"name": "울산본항 UTT부두", "lat": 35.499060, "lon": 129.379200, "radius_m": 520},
    {"name": "울산본항 가스부두", "lat": 35.485900, "lon": 129.385500, "radius_m": 520},
    {"name": "울산본항 용잠1부두", "lat": 35.499009, "lon": 129.375800, "radius_m": 480},
    {"name": "울산본항 용잠2부두", "lat": 35.499009, "lon": 129.375800, "radius_m": 480},
    {"name": "울산본항 남화부두", "lat": 35.477325, "lon": 129.384306, "radius_m": 520},
    {"name": "울산본항 석탄부두", "lat": 35.525300, "lon": 129.382060, "radius_m": 520},
    {"name": "울산본항 양곡부두", "lat": 35.496600, "lon": 129.381000, "radius_m": 520},
    {"name": "울산본항 염포부두", "lat": 35.516500, "lon": 129.395890, "radius_m": 620},
    {"name": "울산본항 일반부두", "lat": 35.507930, "lon": 129.386092, "radius_m": 520},
    {"name": "울산본항 자동차부두", "lat": 35.523100, "lon": 129.391200, "radius_m": 620},
]


def main_berth_zone_from_lat_lon(lat: float, lon: float) -> str | None:
    if lat == 0 or lon == 0:
        return None

    nearest = None
    nearest_km = 999.0

    for area in MAIN_BERTH_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        radius_km = float(area["radius_m"]) / 1000.0
        if distance_km <= radius_km and distance_km < nearest_km:
            nearest = area
            nearest_km = distance_km

    if nearest is None:
        return None

    return str(nearest["name"])


def main_berth_debug_payload(lat: float, lon: float) -> Dict[str, Any]:
    zone = main_berth_zone_from_lat_lon(lat, lon)
    distances = []
    for area in MAIN_BERTH_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        distances.append({
            "name": area["name"],
            "lat": area["lat"],
            "lon": area["lon"],
            "radiusM": area["radius_m"],
            "distanceKm": round(distance_km, 3),
            "inside": distance_km <= float(area["radius_m"]) / 1000.0,
        })
    distances.sort(key=lambda item: item["distanceKm"])
    return {
        "zone": zone,
        "nearest": distances[:5],
    }



# 3.0.8 온산 / 신항 / 미포 / 현대 / KPX / 기타 울산항 부두 좌표 기반 판정
# 2-4순위 작업 범위입니다.
# 기존 모바일 지도에 들어가 있던 부두 기준점을 서버 판정에도 확장 적용합니다.
# 본항 주요 부두는 MAIN_BERTH_AREAS에서 먼저 판정하고,
# 아래 구역은 그 다음 우선순위로 판정합니다.
OTHER_BERTH_AREAS = [
    # 울산신항 / 용연 / 신항 컨테이너 권역
    {"name": "울산신항 LS MNM 신항부두", "lat": 35.436240, "lon": 129.367230, "radius_m": 620},
    {"name": "울산신항 UTK 신항부두", "lat": 35.431880, "lon": 129.370980, "radius_m": 620},
    {"name": "울산신항 대한통운신항부두", "lat": 35.435820, "lon": 129.369530, "radius_m": 620},
    {"name": "울산신항 신항컨부두", "lat": 35.455740, "lon": 129.366700, "radius_m": 620},
    {"name": "울산신항 신항일반부두", "lat": 35.457870, "lon": 129.360860, "radius_m": 620},
    {"name": "울산신항 작업 및 관리부두", "lat": 35.459660, "lon": 129.355010, "radius_m": 620},
    {"name": "울산신항 용연부두", "lat": 35.460970, "lon": 129.371880, "radius_m": 620},
    {"name": "울산신항 신항남방파제 T/S부두", "lat": 35.438750, "lon": 129.378110, "radius_m": 650},
    {"name": "울산신항 신항북방파제 T/S부두", "lat": 35.460690, "lon": 129.384320, "radius_m": 650},

    # 온산항 / OTK / S-OIL / 대한유화 / 달포 권역
    {"name": "온산항 OTK1부두", "lat": 35.461210, "lon": 129.347350, "radius_m": 620},
    {"name": "온산항 OTK2부두", "lat": 35.465350, "lon": 129.345710, "radius_m": 620},
    {"name": "온산항 UTK부두", "lat": 35.469500, "lon": 129.341910, "radius_m": 650},
    {"name": "온산항 S-Oil 1부두", "lat": 35.448310, "lon": 129.354950, "radius_m": 620},
    {"name": "온산항 S-Oil 2부두", "lat": 35.447660, "lon": 129.358450, "radius_m": 620},
    {"name": "온산항 S-Oil 3부두", "lat": 35.451810, "lon": 129.348880, "radius_m": 620},
    {"name": "온산항 S-Oil 4부두", "lat": 35.445420, "lon": 129.358740, "radius_m": 620},
    {"name": "온산항 온산1부두", "lat": 35.438510, "lon": 129.356670, "radius_m": 620},
    {"name": "온산항 온산2부두", "lat": 35.439420, "lon": 129.359470, "radius_m": 620},
    {"name": "온산항 온산3부두", "lat": 35.440460, "lon": 129.362470, "radius_m": 620},
    {"name": "온산항 온산4부두", "lat": 35.441040, "lon": 129.364470, "radius_m": 620},
    {"name": "온산항 달포부두", "lat": 35.439400, "lon": 129.354440, "radius_m": 620},
    {"name": "온산항 대한유화부두", "lat": 35.457020, "lon": 129.347760, "radius_m": 620},

    # 사용자가 요청한 추가 권역. 실제 현장 테스트 후 반경/중심점은 계속 보정할 수 있습니다.
    {"name": "온산항 KPX부두", "lat": 35.453400, "lon": 129.350600, "radius_m": 550},

    # 장생포 / 미포 / 현대 권역
    {"name": "장생포항", "lat": 35.503600, "lon": 129.375500, "radius_m": 850},
    {"name": "미포항 현대미포부두", "lat": 35.524500, "lon": 129.437680, "radius_m": 800},
    {"name": "미포항 미포부두", "lat": 35.524500, "lon": 129.437680, "radius_m": 700},
    {"name": "현대중공업 안벽", "lat": 35.529000, "lon": 129.445000, "radius_m": 1000},
]


def other_berth_zone_from_lat_lon(lat: float, lon: float) -> str | None:
    if lat == 0 or lon == 0:
        return None

    nearest = None
    nearest_km = 999.0

    for area in OTHER_BERTH_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        radius_km = float(area["radius_m"]) / 1000.0
        if distance_km <= radius_km and distance_km < nearest_km:
            nearest = area
            nearest_km = distance_km

    if nearest is None:
        return None

    return str(nearest["name"])


def other_berth_debug_payload(lat: float, lon: float) -> Dict[str, Any]:
    zone = other_berth_zone_from_lat_lon(lat, lon)
    distances = []
    for area in OTHER_BERTH_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        distances.append({
            "name": area["name"],
            "lat": area["lat"],
            "lon": area["lon"],
            "radiusM": area["radius_m"],
            "distanceKm": round(distance_km, 3),
            "inside": distance_km <= float(area["radius_m"]) / 1000.0,
        })
    distances.sort(key=lambda item: item["distanceKm"])
    return {
        "zone": zone,
        "nearest": distances[:8],
    }


def _nearest_point_zone(lat: float, lon: float, areas: List[Dict[str, Any]]) -> Dict[str, Any] | None:
    """
    반경 기반 구역 목록에서 실제 반경 안에 들어온 후보 중 가장 가까운 1개를 고릅니다.
    여러 부두/부이가 겹치더라도 최종 알림 문구에는 이 함수가 고른 1개만 사용합니다.
    """
    if lat == 0 or lon == 0:
        return None

    nearest: Dict[str, Any] | None = None
    nearest_km = 999.0

    for area in areas:
        try:
            area_lat = float(area["lat"])
            area_lon = float(area["lon"])
            radius_km = float(area["radius_m"]) / 1000.0
        except Exception:
            continue

        distance_km = _distance_km(lat, lon, area_lat, area_lon)
        if distance_km <= radius_km and distance_km < nearest_km:
            nearest_km = distance_km
            nearest = dict(area)
            nearest["distanceKm"] = round(distance_km, 3)
            nearest["inside"] = True

    return nearest


def berth_zone_from_lat_lon(lat: float, lon: float) -> str | None:
    """
    본항/온산/신항/미포/현대/KPX 등 모든 부두 후보를 하나로 묶어
    최종 부두명은 거리 기준으로 가장 가까운 1개만 반환합니다.
    """
    berth = _nearest_point_zone(lat, lon, MAIN_BERTH_AREAS + OTHER_BERTH_AREAS)
    if berth is None:
        return None
    return str(berth["name"])


def berth_debug_payload(lat: float, lon: float) -> Dict[str, Any]:
    berth = _nearest_point_zone(lat, lon, MAIN_BERTH_AREAS + OTHER_BERTH_AREAS)
    distances = []
    for area in MAIN_BERTH_AREAS + OTHER_BERTH_AREAS:
        distance_km = _distance_km(lat, lon, float(area["lat"]), float(area["lon"]))
        distances.append({
            "name": area["name"],
            "lat": area["lat"],
            "lon": area["lon"],
            "radiusM": area["radius_m"],
            "distanceKm": round(distance_km, 3),
            "inside": distance_km <= float(area["radius_m"]) / 1000.0,
        })
    distances.sort(key=lambda item: item["distanceKm"])
    return {
        "zone": None if berth is None else str(berth["name"]),
        "selected": berth,
        "nearest": distances[:10],
    }


# 3.1.1 울산항 항로 좌표 기반 판정
# 사용자가 제공한 제1항로/제2항로/제3항로 경계 좌표를 polygon으로 등록합니다.
# 최종 구역 우선순위에서는 묘박지/부이/부두 다음에 항로를 판정합니다.
# 즉, 선박이 부두나 묘박지 안에 있으면 그 구역을 우선 표시하고,
# 그렇지 않은 상태에서 항로 안에 있으면 "제1항로"처럼 표시합니다.
FAIRWAY_AREAS = [
    {
        "name": "제1항로",
        "description": "울산항 진입 메인 항로",
        "coords": [
            (35.40305556, 129.41805556),
            (35.46638889, 129.39422222),
            (35.49116667, 129.39197222),
            (35.51388889, 129.38905556),
            (35.51477778, 129.39230556),
            (35.49269444, 129.39513889),
            (35.46638889, 129.39963889),
            (35.40305556, 129.42416667),
        ],
    },
    {
        "name": "제2항로",
        "description": "울산본항/내항 연결 항로",
        "coords": [
            (35.49116667, 129.39197222),
            (35.49563889, 129.38836111),
            (35.50055556, 129.37883333),
            (35.50222222, 129.37938889),
            (35.49852778, 129.38880556),
            (35.49975000, 129.39086111),
        ],
    },
    {
        "name": "제3항로",
        "description": "온산/신항 방면 연결 항로",
        "coords": [
            (35.43750000, 129.40527778),
            (35.44388889, 129.39638889),
            (35.44822222, 129.37294444),
            (35.45072222, 129.37422222),
            (35.44888889, 129.39555556),
            (35.45166667, 129.39972222),
        ],
    },
]


def fairway_zone_from_lat_lon(lat: float, lon: float) -> str | None:
    if lat == 0 or lon == 0:
        return None

    for area in FAIRWAY_AREAS:
        if point_in_polygon(lat, lon, area["coords"]):
            return str(area["name"])

    return None


def fairway_debug_payload(lat: float, lon: float) -> Dict[str, Any]:
    zone = fairway_zone_from_lat_lon(lat, lon)
    distances = []
    for area in FAIRWAY_AREAS:
        center_lat, center_lon = _polygon_center(area["coords"])
        distances.append({
            "name": area["name"],
            "description": area.get("description", ""),
            "centerLat": round(center_lat, 6),
            "centerLon": round(center_lon, 6),
            "distanceKm": round(_distance_km(lat, lon, center_lat, center_lon), 3),
            "inside": point_in_polygon(lat, lon, area["coords"]),
        })
    distances.sort(key=lambda item: item["distanceKm"])
    return {
        "zone": zone,
        "nearest": distances[:3],
    }


def resolve_zone_decision(lat: float, lon: float) -> Dict[str, Any]:
    """
    3.0.9 좌표 기반 최종 구역 판정 우선순위.

    1순위: M/E 묘박지
    2순위: SK/S-OIL 부이
    3순위: 모든 부두 중 가장 가까운 부두 1개
    4순위: 제1항로/제2항로/제3항로
    5순위: 넓은 해역/접근/항내 구역

    이 함수의 selectedZone이 실제 알림 location/eventType 구역명에 사용됩니다.
    """
    if lat == 0 or lon == 0:
        return {"selectedZone": "위치 확인중", "zoneType": "unknown", "priority": 0}

    anchorage_zone = anchorage_zone_from_lat_lon(lat, lon)
    if anchorage_zone:
        return {"selectedZone": anchorage_zone, "zoneType": "anchorage", "priority": 1}

    buoy_zone = buoy_zone_from_lat_lon(lat, lon)
    if buoy_zone:
        return {"selectedZone": buoy_zone, "zoneType": "buoy", "priority": 2}

    berth_zone = berth_zone_from_lat_lon(lat, lon)
    if berth_zone:
        return {"selectedZone": berth_zone, "zoneType": "berth", "priority": 3}

    fairway_zone = fairway_zone_from_lat_lon(lat, lon)
    if fairway_zone:
        return {"selectedZone": fairway_zone, "zoneType": "route", "priority": 4}

    if lat >= 35.48 and lon >= 129.42:
        return {"selectedZone": "외항/동측 해역", "zoneType": "broad_area", "priority": 5}
    if lat >= 35.43 and lon >= 129.35:
        return {"selectedZone": "울산항 접근 해역", "zoneType": "broad_area", "priority": 5}
    if lat >= 35.37 and lon >= 129.33:
        return {"selectedZone": "울산항 항내/부두권", "zoneType": "broad_area", "priority": 5}
    return {"selectedZone": "울산항 인근", "zoneType": "broad_area", "priority": 5}


def simple_area_from_lat_lon(lat: float, lon: float) -> str:
    decision = resolve_zone_decision(lat, lon)
    return str(decision.get("selectedZone") or "울산항 인근")



def normalize_zone_id(value: str) -> str:
    zone = str(value or "").strip().upper()
    zone = zone.replace(" ", "_").replace("/", "_").replace("-", "_")
    zone = "".join(ch for ch in zone if ch.isalnum() or ch in "_가-힣")
    return zone or "UNKNOWN_ZONE"


def is_berth_area(location: str) -> bool:
    value = str(location or "").upper()
    berth_keywords = [
        "부두", "선석", "접안", "항내/부두권", "부이", "BERTH", "DOCK", "PIER", "WHARF", "BUOY",
        "SK", "S-OIL", "S.OIL", "SOIL", "정일", "UTK", "UTT", "OTK", "본항", "염포", "온산", "용연", "용잠",
        "신항", "미포", "현대", "KPX", "장생포", "대한유화", "대한통운", "달포", "LS", "MNM",
    ]
    return any(keyword.upper() in value for keyword in berth_keywords)


def is_buoy_area(location: str) -> bool:
    """SK/S-OIL 부이처럼 실제 계류 알림을 유지해야 하는 구역인지 판단합니다."""
    value = str(location or "").upper()
    return "부이" in str(location or "") or "BUOY" in value


def is_fairway_area(location: str) -> bool:
    """제1/2/3항로처럼 실제 선박 이동 경로인 항로인지 판단합니다.

    외항/접근해역도 zoneType은 route 계열로 다루지만,
    푸시 알림에서 말하는 '항로 진입'은 제1항로/제2항로/제3항로 진입만 의미합니다.
    """
    value = str(location or "").replace(" ", "")
    return (
        "제1항로" in value
        or "제2항로" in value
        or "제3항로" in value
        or "1항로" in value
        or "2항로" in value
        or "3항로" in value
    )


def is_facility_area(location: str) -> bool:
    """부두/묘박지/부이처럼 선박이 머물 수 있는 시설성 구역인지 판단합니다."""
    return is_berth_area(location) or is_anchorage_area(location)


def fairway_zone_label(location: str) -> str:
    value = str(location or "").strip()
    if not value or value == "위치 확인중":
        return "항로"
    return value


def is_same_zone(a: str, b: str) -> bool:
    return normalize_zone_id(a) == normalize_zone_id(b)


def is_anchorage_area(location: str) -> bool:
    value = str(location or "").upper().replace(" ", "")
    if "묘지" in value or "묘박" in value or "정박지" in value or "ANCHOR" in value:
        return True
    # M1~M7, E1~E3 형태. 3.0.3에서 실제 좌표 기반 세부 구역 판정으로 확장 예정.
    import re
    return re.search(r"(^|[^A-Z0-9])(M[1-7]|E[1-3])([^A-Z0-9]|$)", value) is not None


def berth_zone_label(location: str) -> str:
    value = str(location or "").strip()
    if not value or value == "위치 확인중":
        return "부두권"
    return value


def anchorage_zone_label(location: str) -> str:
    value = str(location or "").strip()
    if not value or value == "위치 확인중":
        return "묘박지"
    return value



def zone_type_from_location(location: str) -> str:
    """
    알림 문구와 이벤트 흐름 정리에 사용할 구역 유형입니다.
    - anchorage: M/E 묘박지
    - berth: 부두/선석/부이
    - route: 항로/접근/항내/외항 같은 넓은 해역
    - unknown: 위치 확인 불가
    """
    value = str(location or "").strip()
    if not value or value == "위치 확인중":
        return "unknown"
    if is_anchorage_area(value):
        return "anchorage"
    if is_berth_area(value):
        return "berth"
    upper = value.upper()
    route_keywords = ["항로", "접근", "항내", "외항", "해역", "PORT", "ROUTE", "APPROACH"]
    if any(keyword in upper for keyword in route_keywords):
        return "route"
    return "broad_area"


def facility_action_words(location: str) -> Dict[str, str]:
    """부두와 부이를 같은 berth 계열로 다루되, 사용자 알림 문구는 자연스럽게 분리합니다."""
    value = str(location or "")
    if "부이" in value or "BUOY" in value.upper():
        return {
            "approaching": "부이 계류중",
            "completed": "부이 계류완료",
            "approach_body": "부이 계류 중",
            "complete_body": "부이 계류완료 상태",
            "flow": "입항 → 부이 접근중 → 계류완료",
        }
    return {
        "approaching": "부두 접안중",
        "completed": "부두 접안완료",
        "approach_body": "접안 중",
        "complete_body": "접안완료 상태",
        "flow": "입항 → 부두 접근중 → 접안완료",
    }


def flow_stage_from_snapshot(location: str, status: str, speed: float) -> Dict[str, str]:
    """
    3.1.0 이벤트 흐름 단계.
    이 값은 푸시 data와 서버 히스토리에 같이 저장되어, 나중에 앱에서 흐름별 필터/표시를 하기 쉽게 만듭니다.
    """
    zone_type = zone_type_from_location(location)
    status_text = str(status or "")
    status_upper = status_text.upper()

    stopped = is_stopped_like(status_text, speed)
    underway = is_underway_like(status_text, speed)

    if zone_type == "anchorage":
        if speed <= 0.1:
            return {"code": "ANCHORAGE_COMPLETED", "label": "투묘완료", "zoneType": zone_type}
        if 0.5 <= speed <= 1.5:
            return {"code": "ANCHORAGE_APPROACHING", "label": "묘지 접근중", "zoneType": zone_type}
        if stopped:
            return {"code": "ANCHORAGE_WAITING", "label": "묘박지 대기", "zoneType": zone_type}
        return {"code": "ANCHORAGE_MOVING", "label": "묘박지 이동", "zoneType": zone_type}

    if zone_type == "berth":
        words = facility_action_words(location)
        berth_completed_status = (
            "MOORED" in status_upper
            or "접안" in status_text
            or "정박" in status_text
            or "BERTH" in status_upper
            or "DOCK" in status_upper
            or "ANCHOR" in status_upper
        )
        if speed <= 0.1 and berth_completed_status:
            return {"code": "BERTH_COMPLETED", "label": words["completed"], "zoneType": zone_type}
        if 0.5 <= speed <= 1.0:
            return {"code": "BERTH_APPROACHING", "label": words["approaching"], "zoneType": zone_type}
        if stopped:
            return {"code": "BERTH_NEAR_STOPPED", "label": "부두권 저속/대기", "zoneType": zone_type}
        return {"code": "BERTH_NEAR_MOVING", "label": "부두권 이동", "zoneType": zone_type}

    if zone_type == "route":
        if underway:
            if "외항" in location or "동측" in location:
                return {"code": "PORT_APPROACH", "label": "입항 접근중", "zoneType": zone_type}
            return {"code": "ROUTE_MOVING", "label": "항로 이동중", "zoneType": zone_type}
        if stopped:
            return {"code": "ROUTE_WAITING", "label": "항로/항내 대기", "zoneType": zone_type}
        return {"code": "ROUTE_CHECK", "label": "항로 확인중", "zoneType": zone_type}

    if underway:
        return {"code": "MOVING", "label": "이동중", "zoneType": zone_type}
    if stopped:
        return {"code": "STOPPED", "label": "정지/대기", "zoneType": zone_type}
    return {"code": "TRACKING", "label": "추적중", "zoneType": zone_type}


def flow_summary_line(current: Dict[str, Any]) -> str:
    stage = str(current.get("flowLabel") or "추적중")
    location = str(current.get("location") or "위치 확인중")
    speed = float(current.get("speed", 0.0) or 0.0)
    return f"📍 구역: {location}\n🧭 흐름: {stage}\n⚡ 속도: {speed:.1f} kn"

def build_ship_snapshot(ship_name: str, info: Dict[str, Any]) -> Dict[str, Any]:
    lat = float(info.get("lat", 0.0) or 0.0)
    lon = float(info.get("lon", 0.0) or 0.0)
    speed = float(info.get("speed", 0.0) or 0.0)
    status = str(info.get("status", "-"))
    location = str(info.get("location") or simple_area_from_lat_lon(lat, lon))
    flow = flow_stage_from_snapshot(location, status, speed)

    return {
        "shipName": normalize_ship_name(ship_name),
        "mmsi": str(info.get("mmsi", "-")),
        "status": status,
        "speed": speed,
        "lat": lat,
        "lon": lon,
        "location": location,
        "zoneType": flow.get("zoneType", "unknown"),
        "flowStage": flow.get("code", "TRACKING"),
        "flowLabel": flow.get("label", "추적중"),
        "destination": str(info.get("destination", "-")),
        "eta": str(info.get("eta", "-")),
        "seenAt": now_iso(),
    }


def detect_ship_events(ship_name: str, current: Dict[str, Any], previous: Dict[str, Any] | None, force: bool = False) -> List[Dict[str, str]]:
    """
    3.1.0 이벤트 흐름 정리.

    목표:
    - 특수 이벤트는 기존처럼 30분 쿨다운 예외, 선박+구역+단계 기준 1회 발송
    - 같은 AIS 갱신에서 특수 이벤트/출항 이벤트가 잡히면 일반 상태변경·위치변경 알림은 생략
    - 알림 문구를 입항→부두접근→접안완료 / 입항→묘지접근→투묘완료 / 정박→출항→항로이동 흐름으로 정리
    - flowStage/zoneType은 current snapshot에 저장되어 이후 앱 표시 확장에 사용할 수 있게 함
    """
    events: List[Dict[str, str]] = []
    name = normalize_ship_name(ship_name)

    current_status = str(current.get("status", "-"))
    current_speed = float(current.get("speed", 0.0) or 0.0)
    current_location = str(current.get("location", "위치 확인중"))
    current_flow_stage = str(current.get("flowStage", "TRACKING"))
    current_flow_label = str(current.get("flowLabel", "추적중"))

    has_previous = isinstance(previous, dict)

    if not has_previous:
        events.append({
            "eventType": "ais_first_detected",
            "title": f"🚢 {name} AIS 최초 감지",
            "body": f"{name} 선박이 울산항 AIS에서 처음 감지되었습니다.\n{flow_summary_line(current)}",
        })

    previous_status = str(previous.get("status", "-")) if has_previous else "-"
    previous_speed = float(previous.get("speed", 0.0) or 0.0) if has_previous else 0.0
    previous_location = str(previous.get("location", "위치 확인중")) if has_previous else "위치 확인중"
    previous_flow_stage = str(previous.get("flowStage", "")) if has_previous else ""

    prev_stopped = is_stopped_like(previous_status, previous_speed)
    now_stopped = is_stopped_like(current_status, current_speed)
    prev_underway = is_underway_like(previous_status, previous_speed)
    now_underway = is_underway_like(current_status, current_speed)

    berth_area = is_berth_area(current_location)
    anchorage_area = is_anchorage_area(current_location)
    berth_zone = berth_zone_label(current_location)
    anchorage_zone = anchorage_zone_label(current_location)
    berth_zone_id = normalize_zone_id(berth_zone)
    anchorage_zone_id = normalize_zone_id(anchorage_zone)

    previous_berth_area = is_berth_area(previous_location) if has_previous else False
    previous_anchorage_area = is_anchorage_area(previous_location) if has_previous else False
    previous_berth_zone = berth_zone_label(previous_location) if has_previous else ""
    previous_anchorage_zone = anchorage_zone_label(previous_location) if has_previous else ""
    same_berth_zone = bool(has_previous and previous_berth_area and berth_area and is_same_zone(previous_berth_zone, berth_zone))
    same_anchorage_zone = bool(has_previous and previous_anchorage_area and anchorage_area and is_same_zone(previous_anchorage_zone, anchorage_zone))
    current_is_buoy = is_buoy_area(berth_zone)

    current_fairway_area = is_fairway_area(current_location)
    previous_fairway_area = is_fairway_area(previous_location) if has_previous else False
    fairway_zone = fairway_zone_label(current_location)
    fairway_zone_id = normalize_zone_id(fairway_zone)
    previous_facility_area = is_facility_area(previous_location) if has_previous else False

    current_status_upper = current_status.upper()
    berth_completed_status = (
        "MOORED" in current_status_upper
        or "접안" in current_status
        or "정박" in current_status
        or "BERTH" in current_status_upper
        or "DOCK" in current_status_upper
        or "ANCHOR" in current_status_upper
    )

    special_event_added = False
    departure_event_added = False
    flow_event_added = False

    # 부두 접근중 알림은 통로를 지나가며 인접 부두가 계속 바뀌는 경우 알림이 과도하게 발생했습니다.
    # 그래서 일반 부두는 "접근중" 푸시를 보내지 않고, 실제로 정지/접안이 확인된 "접안완료"만 보냅니다.
    # 단, SK/S-OIL 부이 계류중은 통로 부두 통과 문제가 상대적으로 적어 기존 접근중 알림을 유지합니다.
    if berth_area and current_is_buoy and 0.5 <= current_speed <= 1.0:
        words = facility_action_words(berth_zone)
        events.append({
            "eventType": f"berth_approaching:{berth_zone_id}",
            "title": f"🚢 {name} {words['approaching']}",
            "body": f"{name} 선박이 {berth_zone} 근처에서 {words['approach_body']}으로 감지되었습니다.\n{flow_summary_line(current)}\n➡️ {words['flow']}",
        })
        special_event_added = True

    # 부두/부이 완료: 부두권 또는 부이권 + 0.1kn 이하 + 정박/접안/MOORED 계열 상태.
    # 일반 부두는 같은 부두 구역이 연속으로 확인된 뒤에만 접안완료를 발송해,
    # SK1로 들어가는 중 SK2/SK3 등 인접 부두를 잠깐 스친 것을 접안으로 오판하지 않게 합니다.
    berth_completed_zone_confirmed = current_is_buoy or same_berth_zone or previous_flow_stage in ("BERTH_NEAR_STOPPED", "BERTH_COMPLETED")
    if berth_area and current_speed <= 0.1 and berth_completed_status and berth_completed_zone_confirmed:
        words = facility_action_words(berth_zone)
        events.append({
            "eventType": f"berth_completed:{berth_zone_id}",
            "title": f"⚓ {name} {words['completed']}",
            "body": f"{name} 선박이 {berth_zone}에서 {words['complete_body']}로 감지되었습니다.\n{flow_summary_line(current)}\n✅ {words['flow']}",
        })
        special_event_added = True

    # 묘박지 투묘중: M/E 묘지 또는 정박지에서 0.5~1.5kn 사이로 감속/진입.
    if anchorage_area and 0.5 <= current_speed <= 1.5:
        events.append({
            "eventType": f"anchorage_approaching:{anchorage_zone_id}",
            "title": f"⚓ {name} 묘지 투묘중",
            "body": f"{name} 선박이 {anchorage_zone}에서 투묘 중으로 감지되었습니다.\n{flow_summary_line(current)}\n➡️ 입항 → 묘지 접근중 → 투묘완료",
        })
        special_event_added = True

    # 묘박지 투묘완료: M/E 묘지 또는 정박지에서 0.1kn 이하.
    if anchorage_area and current_speed <= 0.1:
        events.append({
            "eventType": f"anchorage_completed:{anchorage_zone_id}",
            "title": f"⚓ {name} 묘지 투묘완료",
            "body": f"{name} 선박이 {anchorage_zone}에서 투묘완료 상태로 감지되었습니다.\n{flow_summary_line(current)}\n✅ 입항 → 묘지 접근중 → 투묘완료",
        })
        special_event_added = True

    # 정박/접안/투묘/계류 상태에서 항로로 이동하면 출항 흐름으로 우선 알림.
    # 항로는 선박이 머무는 곳이 아니라 이동 경로이므로, 시설 구역에서 항로로 나오는 변화는 의미 있는 이벤트입니다.
    if has_previous and prev_stopped and now_underway:
        if current_fairway_area and previous_facility_area:
            events.append({
                "eventType": f"departure_to_fairway:{fairway_zone_id}",
                "title": f"🔔 {name} 출항 후 항로 이동",
                "body": f"{name} 선박이 {previous_location}에서 이탈하여 {fairway_zone}로 이동 중입니다.\n📍 현재 구역: {current_location}\n⚡ 현재 속도: {current_speed:.1f} kn",
            })
        else:
            events.append({
                "eventType": "departure_detected",
                "title": f"🔔 {name} 출항 · 항로 이동",
                "body": f"{name} 선박이 {previous_location}에서 이동을 시작했습니다.\n📍 현재 구역: {current_location}\n🧭 흐름: 정박/접안 → 출항 → 항로 이동\n⚡ 현재 속도: {current_speed:.1f} kn",
            })
        departure_event_added = True

    # 외항/접근해역/넓은 해역에서 제1·2·3항로로 들어온 경우에는 항로 진입 알림을 명확하게 발송합니다.
    # 단, 앱 사용자가 이미 항로 위의 선박을 처음 등록한 경우는 baseline 로직에서 기준값만 저장하므로 폭주하지 않습니다.
    fairway_entry_event = (
        has_previous
        and not special_event_added
        and not departure_event_added
        and current_fairway_area
        and not previous_fairway_area
        and current_speed >= 3.0
    )
    if fairway_entry_event:
        events.append({
            "eventType": f"fairway_entered:{fairway_zone_id}",
            "title": f"🧭 {name} {fairway_zone} 진입",
            "body": f"{name} 선박이 울산항 {fairway_zone}로 진입했습니다.\n📍 이전 구역: {previous_location}\n📍 현재 구역: {current_location}\n⚡ 속도: {current_speed:.1f} kn",
        })
        flow_event_added = True

    # 일반 항로/접근 흐름 이벤트.
    # 제1항로→제2항로처럼 항로 안에서의 세부 이동은 푸시를 최소화하기 위해 보내지 않습니다.
    if (
        has_previous
        and not special_event_added
        and not departure_event_added
        and not flow_event_added
        and current_flow_stage in ("PORT_APPROACH", "ROUTE_MOVING")
        and previous_flow_stage != current_flow_stage
        and current_speed >= 3.0
        and not (current_fairway_area and previous_fairway_area)
    ):
        events.append({
            "eventType": f"flow_stage:{current_flow_stage}",
            "title": f"🧭 {name} {current_flow_label}",
            "body": f"{name} 선박의 운항 흐름이 {current_flow_label} 단계로 변경되었습니다.\n📍 현재 구역: {current_location}\n⚡ 속도: {current_speed:.1f} kn",
        })
        flow_event_added = True

    major_event_added = special_event_added or departure_event_added or flow_event_added

    # 특수 이벤트가 없고, 항해→정지로 바뀐 경우에만 포괄 정박/접안 감지.
    if has_previous and prev_underway and now_stopped and not major_event_added:
        events.append({
            "eventType": "anchored_or_docked",
            "title": f"⚓ {name} 정박 · 접안 감지",
            "body": f"{name} 선박이 정박 또는 접안 상태로 감지되었습니다.\n{flow_summary_line(current)}",
        })
        major_event_added = True

    # 같은 갱신에서 큰 흐름 이벤트가 이미 잡혔다면 상태/위치 일반 알림은 생략합니다.
    # 이렇게 해야 '접안완료 + 상태변경 + 위치변경'이 한꺼번에 오는 소음을 줄일 수 있습니다.
    if has_previous and not major_event_added and previous_status != current_status:
        events.append({
            "eventType": f"status_changed:{current_status}",
            "title": f"📡 {name} AIS 상태 변화",
            "body": f"{name} 상태가 {previous_status} → {current_status} 로 변경되었습니다.\n{flow_summary_line(current)}",
        })

    berth_passby_location_change = (
        has_previous
        and previous_berth_area
        and berth_area
        and previous_location != current_location
        and not now_stopped
    )
    berth_entry_passby_location_change = (
        has_previous
        and berth_area
        and not current_is_buoy
        and previous_location != current_location
        and current_speed > 0.2
        and not now_stopped
    )
    fairway_internal_location_change = (
        has_previous
        and previous_fairway_area
        and current_fairway_area
        and previous_location != current_location
        and current_speed >= 1.0
    )
    if (
        has_previous
        and not major_event_added
        and previous_location != current_location
        and current_location != "위치 확인중"
        and not berth_passby_location_change
        and not berth_entry_passby_location_change
        and not fairway_internal_location_change
    ):
        events.append({
            "eventType": f"location_changed:{current_location}",
            "title": f"📍 {name} 위치 변화",
            "body": f"{name} 위치가 {previous_location} → {current_location} 로 변경되었습니다.\n🧭 현재 흐름: {current_flow_label}",
        })

    # 너무 조용한 선박도 force 테스트에서는 감지 이벤트를 확인할 수 있게 합니다.
    if force and not events:
        events.append({
            "eventType": "ais_force_status_check",
            "title": f"🚢 {name} AIS 상태 확인",
            "body": f"{name} 현재 상태: {current_status}, 속도: {current_speed:.1f} kn, 위치: {current_location}, 흐름: {current_flow_label}",
        })

    return events


def perform_ais_check_once(force: bool = False, source: str = "manual") -> Dict[str, Any]:
    """
    UPA AIS를 1회 조회하고, watchlist에 등록된 선박의 상태 변화를 감지해서 해당 사용자에게만 FCM을 발송합니다.
    2.9.13부터는 단순 감지가 아니라 이전 상태와 비교해 입항/출항/정박/상태/위치 변화를 분리합니다.
    """
    if not firebase_ready:
        return {
            "ok": False,
            "error": "firebase not ready",
            "firebaseError": firebase_error,
            "source": source,
            "time": now_iso(),
        }

    watch = read_json(WATCH_FILE, {})
    watched_ships = set()

    for _, item in watch.items():
        ships = unique_ship_list(item.get("ships", [])) if isinstance(item, dict) else []
        watched_ships.update(ships)

    if not watched_ships:
        # 감시 선박이 없는 것은 오류가 아니라 "대기 상태"입니다.
        # 이전 버전은 이 상태를 오류/점검기록으로 누적해서 앱 알림탭을 헷갈리게 만들었습니다.
        return {
            "ok": True,
            "idle": True,
            "error": "no watched ships",
            "message": "감시 선박이 없습니다. 앱에서 추적 선박을 서버와 다시 동기화해야 합니다.",
            "watchedCount": 0,
            "source": source,
            "time": now_iso(),
        }

    prune_baseline_pending_to_watched(watched_ships)

    try:
        ais_list = fetch_upa_ais_list()
    except Exception as exc:
        return {
            "ok": False,
            "error": "ais fetch failed",
            "detail": str(exc),
            "source": source,
            "time": now_iso(),
        }

    detected_ships = {}
    for item in ais_list:
        name = pick_ship_name(item)
        if name and name in watched_ships:
            lat = pick_float(item, ["lat", "latitude", "la", "vslLat"], 0.0)
            lon = pick_float(item, ["lot", "lon", "lng", "longitude", "vslLot"], 0.0)
            detected_ships[name] = {
                "name": name,
                "mmsi": pick_text(item, ["aisMmsi", "mmsi", "MMSI", "mmsiNo"], "-"),
                "status": pick_text(item, ["oprtlSttsKr", "oprtlStts", "status", "navStatus", "nvgtStts", "state"], "-"),
                "speed": pick_float(item, ["sog", "speed", "spd", "knots"], 0.0),
                "lat": lat,
                "lon": lon,
                "location": simple_area_from_lat_lon(lat, lon),
                "destination": pick_text(item, ["destination", "dest", "dstn", "etaDest"], "-"),
                "eta": pick_text(item, ["eta", "etaTime", "arrvTm"], "-"),
            }

    previous_state = read_json(AIS_SHIP_STATE_FILE, {})
    new_state = dict(previous_state) if isinstance(previous_state, dict) else {}

    sent = []
    skipped = []
    errors = []
    event_results = []
    baseline_skipped = []
    baseline_pending = read_json(AIS_BASELINE_FILE, {})
    if not isinstance(baseline_pending, dict):
        baseline_pending = {}
    baseline_pending_changed = False

    for ship_name, info in detected_ships.items():
        target_tokens = tokens_for_ship(ship_name)
        disabled_tokens = notification_disabled_tokens_for_ship(ship_name)
        if not target_tokens and disabled_tokens:
            skipped.append({
                "shipName": ship_name,
                "eventType": "notification_off",
                "reason": "all watchers muted this ship",
                "mutedDeviceCount": len(disabled_tokens),
            })
            # 알림 OFF 선박도 현재 상태 기준값은 저장해야 다음에 알림을 다시 켰을 때 폭주하지 않습니다.
            current_snapshot = build_ship_snapshot(ship_name, info)
            new_state[ship_name] = current_snapshot
            continue
        if not target_tokens:
            continue

        current_snapshot = build_ship_snapshot(ship_name, info)
        previous_snapshot = previous_state.get(ship_name) if isinstance(previous_state, dict) else None
        if not isinstance(previous_snapshot, dict):
            previous_snapshot = None

        baseline_key = normalize_ship_name(ship_name)
        baseline_mode = (
            not force
            and previous_snapshot is None
            and baseline_key in baseline_pending
        )

        if baseline_mode:
            # 첫 동기화 직후에는 현재 상태를 기준값으로 저장만 하고 알림은 보내지 않습니다.
            # 다음 AIS 검사부터 실제 변화가 있을 때만 알림을 보냅니다.
            ship_events = []
            baseline_skipped.append({
                "shipName": ship_name,
                "reason": "initial baseline saved without push",
                "location": str(current_snapshot.get("location", "위치 확인중")),
                "flowStage": str(current_snapshot.get("flowStage", "TRACKING")),
                "flowLabel": str(current_snapshot.get("flowLabel", "추적중")),
            })
            baseline_pending.pop(baseline_key, None)
            baseline_pending_changed = True
        else:
            ship_events = detect_ship_events(ship_name, current_snapshot, previous_snapshot, force=force)

        for event in ship_events:
            event_type = event["eventType"]
            can_send, reason = should_send_ais_alert(ship_name, event_type, force=force)
            if not can_send:
                skipped.append({"shipName": ship_name, "eventType": event_type, "reason": reason})
                continue

            result = send_fcm_to_tokens(
                target_tokens,
                event["title"],
                event["body"],
                {
                    "source": "ulsan_ais_fcm_server",
                    "eventType": event_type,
                    "shipName": ship_name,
                    "status": str(current_snapshot.get("status", "-")),
                    "speed": str(current_snapshot.get("speed", 0.0)),
                    "location": str(current_snapshot.get("location", "위치 확인중")),
                    "zoneType": str(current_snapshot.get("zoneType", "unknown")),
                    "flowStage": str(current_snapshot.get("flowStage", "TRACKING")),
                    "flowLabel": str(current_snapshot.get("flowLabel", "추적중")),
                    "sentAt": now_iso(),
                },
            )

            event_results.append({
                "shipName": ship_name,
                "eventType": event_type,
                "title": event["title"],
                "success": result["success"],
            })

            if result["success"] > 0:
                sent.append({"shipName": ship_name, "eventType": event_type})
            if result["errors"]:
                errors.extend(result["errors"])

        # 감지된 선박은 현재 상태를 저장합니다.
        new_state[ship_name] = current_snapshot

    # 이번 AIS에서 사라진 감시선박도 상태에는 남겨두되, 24시간 이상 오래된 상태는 정리합니다.
    now_dt = datetime.now(timezone.utc)
    cleaned_state = {}
    for ship_name, snapshot in new_state.items():
        if not isinstance(snapshot, dict):
            continue
        seen_at = parse_iso_time(str(snapshot.get("seenAt", "")))
        if seen_at is None or (now_dt - seen_at).total_seconds() <= 24 * 3600:
            cleaned_state[ship_name] = snapshot

    write_json(AIS_SHIP_STATE_FILE, cleaned_state)
    if baseline_pending_changed:
        write_json(AIS_BASELINE_FILE, baseline_pending)

    result_payload = {
        "ok": True,
        "source": source,
        "watchedCount": len(watched_ships),
        "aisCount": len(ais_list),
        "detectedCount": len(detected_ships),
        "detectedShips": sorted(detected_ships.keys()),
        "eventCount": len(event_results),
        "events": event_results,
        "sentShips": sent,
        "skippedShips": skipped,
        "baselineSkippedShips": baseline_skipped,
        "baselineSkippedCount": len(baseline_skipped),
        "baselinePendingCount": safe_len(baseline_pending) if isinstance(baseline_pending, dict) else 0,
        "force": force,
        "cooldownMinutes": AIS_ALERT_COOLDOWN_MINUTES,
        "errors": errors[:5],
        "time": now_iso(),
    }

    history = read_json(EVENT_FILE, [])
    history.insert(0, {
        "type": "check_ais_state_change" if source == "auto" else "check_ais_once_state_change",
        **result_payload,
    })
    write_json(EVENT_FILE, history[:200])

    return result_payload

@app.post("/check-ais-once")
def check_ais_once():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    force = bool(payload.get("force", False))

    started = time.time()
    result = perform_ais_check_once(force=force, source="manual")
    update_health_from_check_result(result, int((time.time() - started) * 1000))
    status = 200 if result.get("ok") or is_no_watched_result(result) else 500
    return jsonify(result), status


@app.get("/auto-check-status")
def auto_check_status():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = build_server_health_payload()
    payload["cooldownMinutes"] = AIS_ALERT_COOLDOWN_MINUTES
    return jsonify(payload)


def auto_checker_loop() -> None:
    global auto_checker_last_run, auto_checker_last_result, auto_checker_run_count
    global auto_checker_last_started, auto_checker_last_completed, auto_checker_last_error
    global auto_checker_consecutive_errors

    print(f"AUTO CHECKER LOOP START enabled={AUTO_CHECK_ENABLED} interval={AUTO_CHECK_INTERVAL_SECONDS}s")

    # 서버 부팅 직후 앱 토큰/감시목록 등록 시간을 조금 기다립니다.
    time.sleep(20)

    while True:
        started_ts = now_iso()
        started_mono = time.time()
        try:
            init_firebase()
            with health_lock:
                auto_checker_last_started = started_ts
                auto_checker_last_run = started_ts
                auto_checker_run_count += 1

            result = perform_ais_check_once(force=False, source="auto")
            auto_checker_last_result = result
            update_health_from_check_result(result, int((time.time() - started_mono) * 1000))
            print(f"AUTO CHECK RESULT: {result}")
        except Exception as exc:
            error_payload = {
                "ok": False,
                "error": str(exc),
                "source": "auto",
                "time": now_iso(),
            }
            with health_lock:
                auto_checker_last_run = started_ts
                auto_checker_last_completed = now_iso()
                auto_checker_last_error = str(exc)
                auto_checker_consecutive_errors += 1
                auto_checker_last_result = error_payload
            print(f"AUTO CHECK ERROR: {exc}")

        time.sleep(max(60, AUTO_CHECK_INTERVAL_SECONDS))

def ensure_auto_checker_running() -> None:
    """
    Gunicorn/Render 환경에서는 앱 import 시점과 실제 worker 실행 시점이 달라질 수 있습니다.
    그래서 요청이 들어올 때마다 현재 프로세스(pid) 안에서 자동감시 스레드가 살아있도록 보장합니다.
    3.1.2부터는 thread.is_alive()까지 확인해서 스레드가 죽었으면 다시 시작합니다.
    """
    global auto_checker_started, auto_checker_pid, auto_checker_thread

    if not AUTO_CHECK_ENABLED:
        return

    current_pid = os.getpid()

    if (
        auto_checker_started
        and auto_checker_pid == current_pid
        and auto_checker_thread is not None
        and getattr(auto_checker_thread, "is_alive", lambda: False)()
    ):
        return

    auto_checker_started = True
    auto_checker_pid = current_pid

    thread = threading.Thread(target=auto_checker_loop, daemon=True)
    auto_checker_thread = thread
    thread.start()
    print(f"AUTO CHECKER THREAD STARTED pid={current_pid}")

def start_auto_checker() -> None:
    ensure_auto_checker_running()


@app.post("/register-watch")
def register_watch():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    token = str(payload.get("token", "")).strip()
    ships = payload.get("ships", [])
    notification_off_ships = notification_off_ship_list_from_payload(payload)

    if not token:
        return jsonify({"ok": False, "error": "token is required"}), 400
    if not isinstance(ships, list):
        return jsonify({"ok": False, "error": "ships must be a list"}), 400

    normalized_ships = unique_ship_list(ships)
    notification_off_ships = [name for name in notification_off_ships if name in normalized_ships]
    watch = read_json(WATCH_FILE, {})
    watch[token] = {
        "ships": normalized_ships,
        "notificationOffShips": notification_off_ships,
        "updatedAt": now_iso(),
    }
    write_json(WATCH_FILE, watch)

    baseline_pending_added = mark_baseline_pending_for_missing_state(
        normalized_ships,
        source="register-watch",
        token=token,
    )

    ship_name = normalized_ships[-1] if normalized_ships else ""
    summary = watch_summary(watch, token=token, ship_name=ship_name)

    print(
        f"REGISTER WATCH token={token[:12]}... ships={normalized_ships} "
        f"baselinePending={baseline_pending_added} summary={summary}"
    )

    return jsonify({
        "ok": True,
        "registered": True,
        "registeredShips": len(normalized_ships),
        "ships": normalized_ships,
        "notificationOffShips": notification_off_ships,
        "notificationOffCount": len(notification_off_ships),
        "baselinePendingAdded": baseline_pending_added,
        "baselinePendingAddedCount": len(baseline_pending_added),
        **summary,
        "time": now_iso(),
    })


@app.post("/set-ship-notification")
def set_ship_notification():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    token = str(payload.get("token", "")).strip()
    ship_name = normalize_ship_name(payload.get("shipName", ""))
    enabled = payload.get("enabled", True)
    enabled_bool = bool(enabled)

    if not token:
        return jsonify({"ok": False, "error": "token is required"}), 400
    if not ship_name:
        return jsonify({"ok": False, "error": "shipName is required"}), 400

    watch = read_json(WATCH_FILE, {})
    item = watch.get(token) if isinstance(watch, dict) else None
    if not isinstance(item, dict):
        item = {"ships": [], "updatedAt": now_iso()}

    ships = unique_ship_list(item.get("ships", []))
    if ship_name not in ships:
        ships.append(ship_name)

    notification_off_ships = unique_ship_list(item.get("notificationOffShips", []))
    if enabled_bool:
        notification_off_ships = [name for name in notification_off_ships if name != ship_name]
    elif ship_name not in notification_off_ships:
        notification_off_ships.append(ship_name)

    item["ships"] = ships
    item["notificationOffShips"] = [name for name in notification_off_ships if name in ships]
    item["updatedAt"] = now_iso()
    item["source"] = "set-ship-notification"
    watch[token] = item
    write_json(WATCH_FILE, watch)

    summary = watch_summary(watch, token=token, ship_name=ship_name)
    print(f"SET SHIP NOTIFICATION token={token[:12]}... ship={ship_name} enabled={enabled_bool} summary={summary}")

    return jsonify({
        "ok": True,
        "shipName": ship_name,
        "notificationEnabled": enabled_bool,
        "notificationOffShips": item["notificationOffShips"],
        "notificationOffCount": len(item["notificationOffShips"]),
        **summary,
        "time": now_iso(),
    })


@app.post("/unregister-watch")
def unregister_watch():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    token = str(payload.get("token", "")).strip()
    ships = payload.get("ships", [])
    notification_off_ships = notification_off_ship_list_from_payload(payload)

    if not token:
        return jsonify({"ok": False, "error": "token is required"}), 400
    if not isinstance(ships, list):
        return jsonify({"ok": False, "error": "ships must be a list"}), 400

    normalized_ships = unique_ship_list(ships)
    notification_off_ships = [name for name in notification_off_ships if name in normalized_ships]
    watch = read_json(WATCH_FILE, {})

    if normalized_ships:
        watch[token] = {
            "ships": normalized_ships,
            "notificationOffShips": notification_off_ships,
            "updatedAt": now_iso(),
        }
    else:
        watch.pop(token, None)

    write_json(WATCH_FILE, watch)

    ship_name = str(payload.get("shipName", "")).strip()
    summary = watch_summary(watch, token=token, ship_name=ship_name)

    print(f"UNREGISTER WATCH token={token[:12]}... ships={normalized_ships} summary={summary}")

    return jsonify({
        "ok": True,
        "unregistered": True,
        "registeredShips": len(normalized_ships),
        "ships": normalized_ships,
        "notificationOffShips": notification_off_ships,
        "notificationOffCount": len(notification_off_ships),
        **summary,
        "time": now_iso(),
    })


@app.post("/watch-status")
def watch_status():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    token = str(payload.get("token", "")).strip()

    watch = read_json(WATCH_FILE, {})
    my_ships = []
    my_notification_off_ships = []
    if token and token in watch and isinstance(watch[token], dict):
        my_ships = unique_ship_list(watch[token].get("ships", []))
        my_notification_off_ships = unique_ship_list(watch[token].get("notificationOffShips", []))

    summary = watch_summary(watch, token=token)
    return jsonify({
        "ok": True,
        "ships": my_ships,
        "notificationOffShips": my_notification_off_ships,
        "notificationOffCount": len(my_notification_off_ships),
        **summary,
        "time": now_iso(),
    })


def extract_ship_list_from_payload(payload: Dict[str, Any]) -> List[str]:
    """
    앱 버전에 따라 감시 선박 목록 키 이름이 달라질 수 있어 여러 이름을 허용합니다.
    예: ships, shipNames, watchShips, trackedShips
    """
    for key in ("ships", "shipNames", "watchShips", "trackedShips", "watchList"):
        values = payload.get(key)
        if isinstance(values, list):
            return unique_ship_list(values)
    return []


@app.route("/my-watch", methods=["GET", "POST"])
def my_watch():
    """
    모바일 앱 호환용 감시목록 엔드포인트입니다.

    현재 앱 로그에서 POST /my-watch 호출이 확인되었는데, 기존 서버에는 이 주소가 없어 404가 발생했습니다.
    이 엔드포인트는 다음 두 역할을 동시에 처리합니다.
    1) token만 오면 서버에 저장된 내 감시목록 조회
    2) token + ships 계열 목록이 같이 오면 서버 감시목록을 복구/동기화
    """
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    payload = request.get_json(silent=True) or {}
    token = (
        str(payload.get("token", "")).strip()
        or str(request.args.get("token", "")).strip()
    )

    watch = read_json(WATCH_FILE, {})
    if not isinstance(watch, dict):
        watch = {}

    incoming_ships = extract_ship_list_from_payload(payload)
    incoming_notification_off_ships = notification_off_ship_list_from_payload(payload)
    synced = False

    baseline_pending_added: List[str] = []

    # Render 재배포 후 /tmp 데이터가 비어도 앱이 로컬 SharedPreferences의 선박목록을 보내주면 즉시 복구합니다.
    if token and incoming_ships:
        incoming_notification_off_ships = [name for name in incoming_notification_off_ships if name in incoming_ships]
        watch[token] = {
            "ships": incoming_ships,
            "notificationOffShips": incoming_notification_off_ships,
            "updatedAt": now_iso(),
            "source": "my-watch-sync",
        }
        write_json(WATCH_FILE, watch)
        baseline_pending_added = mark_baseline_pending_for_missing_state(
            incoming_ships,
            source="my-watch-sync",
            token=token,
        )
        synced = True

    my_ships = []
    my_notification_off_ships = []
    if token and token in watch and isinstance(watch[token], dict):
        my_ships = unique_ship_list(watch[token].get("ships", []))
        my_notification_off_ships = unique_ship_list(watch[token].get("notificationOffShips", []))

    summary = watch_summary(watch, token=token)

    print(
        f"MY WATCH token={token[:12]}... incoming={incoming_ships} "
        f"synced={synced} baselinePending={baseline_pending_added} "
        f"myShips={my_ships} notificationOff={my_notification_off_ships} summary={summary}"
    )

    return jsonify({
        "ok": True,
        "synced": synced,
        "registered": synced,
        "ships": my_ships,
        "notificationOffShips": my_notification_off_ships,
        "notificationOffCount": len(my_notification_off_ships),
        "registeredShips": len(my_ships),
        "baselinePendingAdded": baseline_pending_added,
        "baselinePendingAddedCount": len(baseline_pending_added),
        **summary,
        "time": now_iso(),
    })


@app.get("/ais-state")
def ais_state():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    state = read_json(AIS_SHIP_STATE_FILE, {})
    return jsonify({
        "ok": True,
        "stateCount": len(state) if isinstance(state, dict) else 0,
        "state": state if isinstance(state, dict) else {},
        "time": now_iso(),
    })


@app.get("/alert-history")
def alert_history():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    history = read_json(EVENT_FILE, [])
    alert_state = read_json(AIS_ALERT_FILE, {})

    return jsonify({
        "ok": True,
        "historyCount": len(history),
        "recentHistory": history[:30] if isinstance(history, list) else [],
        "cooldownCount": len(alert_state) if isinstance(alert_state, dict) else 0,
        "cooldownState": alert_state if isinstance(alert_state, dict) else {},
        "cooldownMinutes": AIS_ALERT_COOLDOWN_MINUTES,
        "time": now_iso(),
    })


@app.post("/clear-alert-history")
def clear_alert_history():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    write_json(EVENT_FILE, [])
    write_json(AIS_ALERT_FILE, {})
    write_json(AIS_SHIP_STATE_FILE, {})
    write_json(AIS_BASELINE_FILE, {})

    return jsonify({
        "ok": True,
        "cleared": True,
        "message": "alert history, duplicate state, ais ship state, baseline pending cleared",
        "time": now_iso(),
    })


@app.post("/portmis/upload-excel")
def portmis_upload_excel():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    if load_workbook is None:
        return jsonify({
            "ok": False,
            "error": "openpyxl not installed",
            "message": "requirements.txt에 openpyxl==3.1.5 를 추가한 뒤 Render를 재배포하세요.",
        }), 500

    upload = request.files.get("file") or request.files.get("excel") or request.files.get("xlsx")
    if upload is None:
        return jsonify({
            "ok": False,
            "error": "excel file is required",
            "message": "multipart/form-data 형식으로 file 필드에 Port-MIS download.xlsx를 업로드하세요.",
        }), 400

    filename = str(getattr(upload, "filename", "") or "download.xlsx")
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return jsonify({
            "ok": False,
            "error": "invalid file type",
            "filename": filename,
            "message": "Port-MIS 엑셀 .xlsx 파일만 업로드하세요.",
        }), 400

    try:
        parsed = parse_portmis_excel_file(upload)
        parsed["uploadedFilename"] = filename

        previous = read_json(PORTMIS_FILE, {})
        if isinstance(previous, dict) and previous:
            write_json(PORTMIS_BACKUP_FILE, previous)

        write_json(PORTMIS_FILE, parsed)

        return jsonify({
            "ok": True,
            "message": "Port-MIS 엑셀 업로드/파싱 완료",
            "source": parsed.get("source"),
            "from": parsed.get("from"),
            "to": parsed.get("to"),
            "uploadedAt": parsed.get("uploadedAt"),
            "count": parsed.get("count"),
            "portCounts": parsed.get("portCounts", {}),
            "movementCounts": parsed.get("movementCounts", {}),
            "etaPolicy": parsed.get("etaPolicy", {}),
            "sample": parsed.get("items", [])[:3],
            "time": now_iso(),
        })
    except Exception as exc:
        return jsonify({
            "ok": False,
            "error": "failed to parse Port-MIS excel",
            "detail": str(exc),
            "time": now_iso(),
        }), 500


@app.get("/portmis/status")
def portmis_status():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401
    return jsonify(portmis_status_payload(include_items=False))


@app.get("/portmis/weekly")
def portmis_weekly():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401
    return jsonify(portmis_filtered_items(request.args))


@app.get("/portmis/ships")
def portmis_ships():
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    result = portmis_filtered_items(request.args)
    if not result.get("available"):
        return jsonify(result)

    ships = []
    seen = set()
    for item in result.get("items", []):
        key = str(item.get("shipMatchKey") or "")
        if not key or key in seen:
            continue
        seen.add(key)
        ships.append({
            "shipName": item.get("shipName", ""),
            "shipNameRaw": item.get("shipNameRaw", ""),
            "callSign": item.get("callSign", ""),
            "portName": item.get("portName", ""),
            "movementType": item.get("movementType", ""),
            "arrivalTime": item.get("arrivalTime", ""),
            "departureTime": item.get("departureTime", ""),
            "portmisEta": item.get("portmisEta", ""),
            "eta": item.get("eta", ""),
            "etaSource": item.get("etaSource", ""),
            "etaPriority": item.get("etaPriority", 0),
            "berthName": item.get("berthName", ""),
            "nextPort": item.get("nextPort", ""),
            "previousPort": item.get("previousPort", ""),
            "shipType": item.get("shipType", ""),
            "pilotYn": item.get("pilotYn", ""),
        })

    return jsonify({
        "ok": True,
        "available": True,
        "source": result.get("source"),
        "uploadedAt": result.get("uploadedAt"),
        "from": result.get("from"),
        "to": result.get("to"),
        "count": len(ships),
        "etaPolicy": result.get("etaPolicy", {}),
        "ships": ships,
        "time": now_iso(),
    })



@app.get("/zone-test")
def zone_test():
    """
    서버 좌표 판정 테스트용 엔드포인트.
    예: /zone-test?lat=35.46638889&lon=129.39422222
    """
    if not require_api_key():
        return jsonify({"ok": False, "error": "invalid api key"}), 401

    lat = request.args.get("lat", "").strip()
    lon = request.args.get("lon", "").strip()
    try:
        lat_f = float(lat)
        lon_f = float(lon)
    except Exception:
        return jsonify({
            "ok": False,
            "error": "lat and lon query parameters are required",
            "example": "/zone-test?lat=35.46638889&lon=129.39422222",
            "time": now_iso(),
        }), 400

    anchorage_debug = anchorage_debug_payload(lat_f, lon_f)
    buoy_debug = buoy_debug_payload(lat_f, lon_f)
    main_berth_debug = main_berth_debug_payload(lat_f, lon_f)
    other_berth_debug = other_berth_debug_payload(lat_f, lon_f)
    berth_debug = berth_debug_payload(lat_f, lon_f)
    fairway_debug = fairway_debug_payload(lat_f, lon_f)
    decision = resolve_zone_decision(lat_f, lon_f)
    return jsonify({
        "ok": True,
        "lat": lat_f,
        "lon": lon_f,
        "area": decision["selectedZone"],
        "selectedZone": decision["selectedZone"],
        "selectedZoneType": decision["zoneType"],
        "selectedPriority": decision["priority"],
        "priorityRule": "1 anchorage > 2 buoy > 3 nearest berth > 4 fairway route > 5 broad area",
        "anchorageZone": anchorage_debug["zone"],
        "nearestAnchorages": anchorage_debug["nearest"],
        "buoyZone": buoy_debug["zone"],
        "nearestBuoys": buoy_debug["nearest"],
        "berthZone": berth_debug["zone"],
        "nearestBerths": berth_debug["nearest"],
        "selectedBerth": berth_debug["selected"],
        "fairwayZone": fairway_debug["zone"],
        "nearestFairways": fairway_debug["nearest"],
        "mainBerthZone": main_berth_debug["zone"],
        "nearestMainBerths": main_berth_debug["nearest"],
        "otherBerthZone": other_berth_debug["zone"],
        "nearestOtherBerths": other_berth_debug["nearest"],
        "time": now_iso(),
    })


@app.get("/alerts/demo")
def alerts_demo():
    return jsonify({
        "ok": True,
        "alerts": [
            {
                "id": f"demo-{int(time.time() // 60)}",
                "type": "🌐",
                "title": "FCM 서버 연결 테스트",
                "body": "Render FCM 서버가 정상 응답했습니다.",
            }
        ]
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port)
